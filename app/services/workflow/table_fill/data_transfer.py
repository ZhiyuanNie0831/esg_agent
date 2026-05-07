"""Excel 源表到目标表直填导出。

这个模块处理“不需要计算，只把源表数据写入目标表”的场景。
它会优先识别目标表中已经存在的 key 行并回填空白单元格；
没有稳定 key 行时，才按目标表表头追加空白行，并保留逐单元格审计。
"""

from __future__ import annotations

import base64
import logging
import os
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.workflow.table_fill.crosscheck import TableDataTransferCrossCheckService
from app.services.workflow.uploads.store import workflow_upload_store

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
RESULT_SHEET_TITLE = "数据写入结果"
PLACEHOLDER_CELL_VALUES = {"", "-", "--", "待填", "待填写", "待补充", "tbd", "n/a", "na", "pending"}
HEADER_SCAN_ROW_LIMIT = 8
logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class SourceTable:
    document: Any
    sheet: dict[str, Any]
    headers: list[str]
    rows: list[dict[str, Any]]


@dataclass(frozen=True)
class TargetSheet:
    worksheet: Worksheet
    header_row_index: int
    headers: list[str]


class TableDataTransferExporter:
    """把源 Excel 中的明细行按目标表头写入目标 workbook。"""

    def __init__(self, crosscheck_service: TableDataTransferCrossCheckService | None = None) -> None:
        self._crosscheck_service = crosscheck_service or TableDataTransferCrossCheckService()

    def build_export_bundle(
        self,
        *,
        task: str,
        source_documents: list[Any],
        target_documents: list[Any],
    ) -> dict[str, object]:
        source_table = self._select_source_table(source_documents)
        if source_table is None:
            return {
                "summary": "没有找到可写入的源数据表。",
                "exportFiles": [],
                "transferAudit": [],
                "transferStats": {"mode": "no_source", "written": 0, "rowsTransferred": 0},
            }

        workbook, target_sheet_names, used_original_workbook = self._build_target_workbook(target_documents)
        target_sheet = self._select_target_sheet(
            workbook=workbook,
            target_sheet_names=target_sheet_names,
            source_headers=source_table.headers,
            task=task,
        )
        if target_sheet is None:
            target_sheet = self._create_result_sheet(workbook, source_table.headers)
            mode = "result_sheet"
        else:
            mode = "header_mapping"

        column_mappings = self._match_columns(
            source_headers=source_table.headers,
            target_headers=target_sheet.headers,
        )
        if not column_mappings:
            target_sheet = self._create_result_sheet(workbook, source_table.headers)
            column_mappings = [
                {
                    "sourceHeader": header,
                    "targetHeader": header,
                    "targetColumnIndex": index,
                    "score": 1.0,
                    "reason": "未识别到目标表头映射，已回退为完整结果表。",
                }
                for index, header in enumerate(source_table.headers, start=1)
            ]
            mode = "result_sheet"

        crosscheck = self._crosscheck_service.review_transfer_plan(
            task=task,
            source_summary={
                "workbook": getattr(source_table.document, "name", ""),
                "sheet": source_table.sheet.get("title"),
                "headers": source_table.headers,
                "rowCount": len(source_table.rows),
                "sampleRows": source_table.rows[:3],
            },
            target_summary={
                "workbook": self._build_export_filename(target_documents),
                "sheet": target_sheet.worksheet.title,
                "headers": target_sheet.headers,
                "headerRowIndex": target_sheet.header_row_index,
            },
            column_mappings=column_mappings,
            row_count=len(source_table.rows),
        )
        if bool(crosscheck.get("blockWrite")):
            return {
                "summary": "Agent 交叉检查认为当前写入计划风险较高，已阻止自动写入。请人工确认源表、目标表和列映射后再执行。",
                "sourceWorkbook": getattr(source_table.document, "name", ""),
                "sourceSheet": source_table.sheet.get("title"),
                "targetWorkbook": self._build_export_filename(target_documents),
                "targetSheet": target_sheet.worksheet.title,
                "columnMappings": column_mappings,
                "crossCheck": crosscheck,
                "transferAudit": [
                    {
                        "status": "blocked_by_crosscheck",
                        "mode": mode,
                        "sourceDocument": getattr(source_table.document, "name", ""),
                        "sourceSheet": source_table.sheet.get("title"),
                        "targetSheet": target_sheet.worksheet.title,
                        "decisionSource": "agent_review",
                        "riskLevel": crosscheck.get("riskLevel"),
                        "reasonSummary": "；".join(str(item) for item in crosscheck.get("issues", []) or []),
                        "message": "Agent 交叉检查阻止了本次自动写入。",
                    }
                ],
                "transferStats": {
                    "mode": mode,
                    "written": 0,
                    "keptExisting": 0,
                    "preservedExisting": 0,
                    "sourceEmpty": 0,
                    "rowsTransferred": 0,
                    "blockedByCrossCheck": 1,
                },
                "exportFiles": [],
            }

        keyed_fill_plan = self._build_keyed_fill_plan(
            source_table=source_table,
            target_sheet=target_sheet,
            column_mappings=column_mappings,
        )
        if keyed_fill_plan:
            mode = "keyed_blank_fill"
            audit = self._write_keyed_rows(
                source_table=source_table,
                target_sheet=target_sheet,
                fill_plan=keyed_fill_plan,
                mode=mode,
            )
        else:
            audit = self._write_rows(
                source_table=source_table,
                target_sheet=target_sheet,
                column_mappings=column_mappings,
                mode=mode,
            )
        stats = self._summarize_audit(audit, mode=mode)
        target_filename = self._build_export_filename(target_documents)
        summary = (
            f"已将源表“{getattr(source_table.document, 'name', '')} / {source_table.sheet.get('title')}”"
            f"中的 {stats['rowsTransferred']} 行数据写入目标工作表“{target_sheet.worksheet.title}”。"
            f" 写入 {stats['written']} 个单元格。"
        )
        if not used_original_workbook:
            summary += " 当前导出基于解析后的结构化数据重建目标 workbook。"

        return {
            "summary": summary,
            "sourceWorkbook": getattr(source_table.document, "name", ""),
            "sourceSheet": source_table.sheet.get("title"),
            "targetWorkbook": target_filename,
            "targetSheet": target_sheet.worksheet.title,
            "columnMappings": column_mappings,
            "crossCheck": crosscheck,
            "transferAudit": audit,
            "transferStats": stats,
            "exportFiles": [
                {
                    "label": f"已自动写入目标表“{target_sheet.worksheet.title}”。",
                    "filename": target_filename,
                    "mimeType": XLSX_MIME_TYPE,
                    "contentBase64": self._encode_workbook(workbook),
                }
            ],
        }

    def _select_source_table(self, source_documents: list[Any]) -> SourceTable | None:
        candidates: list[tuple[int, int, Any, dict[str, Any]]] = []
        for document_index, document in enumerate(source_documents):
            if getattr(document, "type", None) != "excel":
                continue
            for sheet in getattr(document, "structuredData", {}).get("sheets", []) or []:
                if not isinstance(sheet, dict):
                    continue
                headers = [str(item) for item in sheet.get("headers", []) or [] if str(item).strip()]
                rows = [item for item in sheet.get("rows", []) or [] if isinstance(item, dict)]
                if not headers or not rows:
                    continue
                score = len(rows) + len(sheet.get("numericColumns", []) or []) * 2
                if str(sheet.get("sheetRole") or "") == "source_like":
                    score += 5
                candidates.append((score, -document_index, document, sheet))

        if not candidates:
            return None

        candidates.sort(reverse=True, key=lambda item: (item[0], item[1]))
        _score, _order, document, sheet = candidates[0]
        headers = [str(item) for item in sheet.get("headers", []) or [] if str(item).strip()]
        rows = [dict(item) for item in sheet.get("rows", []) or [] if isinstance(item, dict)]
        return SourceTable(document=document, sheet=sheet, headers=headers, rows=rows)

    def _build_target_workbook(self, target_documents: list[Any]) -> tuple[Workbook, list[str], bool]:
        original = self._load_original_workbook(target_documents)
        if original is not None:
            return original, list(original.sheetnames), True

        workbook = Workbook()
        workbook.remove(workbook.active)
        sheet_names: list[str] = []
        for document in target_documents:
            if getattr(document, "type", None) != "excel":
                continue
            for sheet in getattr(document, "structuredData", {}).get("sheets", []) or []:
                if not isinstance(sheet, dict):
                    continue
                title = self._make_unique_sheet_name(workbook, str(sheet.get("title") or "Sheet"))
                worksheet = workbook.create_sheet(title=title)
                sheet_names.append(title)
                self._write_structured_sheet(worksheet, sheet)

        if not workbook.worksheets:
            worksheet = workbook.create_sheet(title=RESULT_SHEET_TITLE)
            sheet_names.append(worksheet.title)

        return workbook, sheet_names, False

    def _load_original_workbook(self, target_documents: list[Any]) -> Workbook | None:
        excel_documents = [document for document in target_documents if getattr(document, "type", None) == "excel"]
        if len(excel_documents) != 1:
            return None
        token = str(excel_documents[0].structuredData.get("workbookUploadToken") or "").strip()
        workbook_bytes = workflow_upload_store.get_bytes(token)
        if not workbook_bytes:
            return None
        try:
            return load_workbook(BytesIO(workbook_bytes))
        except Exception:
            logger.exception(
                "table data transfer original workbook load failed; falling back to structured workbook: document=%s",
                getattr(excel_documents[0], "name", ""),
            )
            return None

    def _write_structured_sheet(self, worksheet: Worksheet, sheet: dict[str, Any]) -> None:
        headers = [str(header) for header in sheet.get("headers", []) or [] if str(header).strip()]
        if headers:
            worksheet.append(headers)
            for row in sheet.get("rows", []) or []:
                if isinstance(row, dict):
                    worksheet.append([row.get(header) for header in headers])

        for merged_range in sheet.get("mergedRanges", []) or []:
            range_ref = str((merged_range or {}).get("range") if isinstance(merged_range, dict) else merged_range).strip()
            if range_ref:
                worksheet.merge_cells(range_ref)

    def _select_target_sheet(
        self,
        *,
        workbook: Workbook,
        target_sheet_names: list[str],
        source_headers: list[str],
        task: str,
    ) -> TargetSheet | None:
        candidates: list[tuple[float, str, TargetSheet]] = []
        allowed_sheet_names = set(target_sheet_names) or set(workbook.sheetnames)
        task_lower = str(task or "").lower()

        for sheet_name in workbook.sheetnames:
            if sheet_name not in allowed_sheet_names:
                continue
            worksheet = workbook[sheet_name]
            target_sheet = self._detect_target_header(worksheet)
            if target_sheet is None:
                continue
            score = self._score_target_sheet(target_sheet, source_headers)
            if sheet_name.lower() in task_lower:
                score += 0.25
            if score > 0:
                candidates.append((score, sheet_name, target_sheet))

        if not candidates:
            return None

        candidates.sort(key=lambda item: (-item[0], item[1]))
        return candidates[0][2]

    def _detect_target_header(self, worksheet: Worksheet) -> TargetSheet | None:
        scan_limit = min(max(worksheet.max_row, 1), HEADER_SCAN_ROW_LIMIT)
        best: tuple[int, int, list[str]] | None = None
        for row_index in range(1, scan_limit + 1):
            values = [worksheet.cell(row=row_index, column=column_index).value for column_index in range(1, worksheet.max_column + 1)]
            headers = [str(value).strip() if value not in (None, "") else "" for value in values]
            non_empty_count = len([header for header in headers if header])
            if non_empty_count < 2:
                continue
            score = non_empty_count
            if best is None or score > best[0]:
                best = (score, row_index, headers)

        if best is None:
            return None
        _score, row_index, headers = best
        return TargetSheet(worksheet=worksheet, header_row_index=row_index, headers=headers)

    def _score_target_sheet(self, target_sheet: TargetSheet, source_headers: list[str]) -> float:
        mappings = self._match_columns(source_headers=source_headers, target_headers=target_sheet.headers)
        if not mappings:
            return 0.0
        return sum(float(item["score"]) for item in mappings) / max(len(source_headers), 1)

    def _create_result_sheet(self, workbook: Workbook, headers: list[str]) -> TargetSheet:
        if RESULT_SHEET_TITLE in workbook.sheetnames:
            workbook.remove(workbook[RESULT_SHEET_TITLE])
        worksheet = workbook.create_sheet(title=RESULT_SHEET_TITLE)
        worksheet.append(headers)
        return TargetSheet(worksheet=worksheet, header_row_index=1, headers=list(headers))

    def _match_columns(
        self,
        *,
        source_headers: list[str],
        target_headers: list[str],
    ) -> list[dict[str, object]]:
        mappings: list[dict[str, object]] = []
        used_source_headers: set[str] = set()
        normalized_sources = [
            (header, self._normalize_header(header))
            for header in source_headers
            if str(header).strip()
        ]

        for target_index, target_header in enumerate(target_headers, start=1):
            normalized_target = self._normalize_header(target_header)
            if not normalized_target:
                continue
            scored_sources: list[tuple[float, str, str]] = []
            for source_header, normalized_source in normalized_sources:
                if source_header in used_source_headers:
                    continue
                score = self._score_header_match(normalized_source, normalized_target)
                if score <= 0:
                    continue
                scored_sources.append((score, source_header, normalized_source))
            if not scored_sources:
                continue
            scored_sources.sort(key=lambda item: (-item[0], item[1]))
            score, source_header, _normalized_source = scored_sources[0]
            used_source_headers.add(source_header)
            mappings.append(
                {
                    "sourceHeader": source_header,
                    "targetHeader": str(target_header).strip(),
                    "targetColumnIndex": target_index,
                    "score": round(score, 3),
                    "reason": "表头完全匹配。" if score >= 1.0 else "表头相似，已自动匹配。",
                }
            )

        return mappings

    def _score_header_match(self, normalized_source: str, normalized_target: str) -> float:
        if not normalized_source or not normalized_target:
            return 0.0
        if normalized_source == normalized_target:
            return 1.0
        if len(normalized_source) >= 2 and len(normalized_target) >= 2:
            if normalized_source in normalized_target or normalized_target in normalized_source:
                return 0.72
        return 0.0

    def _build_keyed_fill_plan(
        self,
        *,
        source_table: SourceTable,
        target_sheet: TargetSheet,
        column_mappings: list[dict[str, object]],
    ) -> dict[str, object] | None:
        key_mapping = self._select_key_mapping(column_mappings)
        if key_mapping is None:
            return None

        source_key_header = str(key_mapping["sourceHeader"])
        target_key_column = int(key_mapping["targetColumnIndex"])
        target_rows_by_key = self._index_target_rows_by_key(
            target_sheet=target_sheet,
            target_key_column=target_key_column,
        )
        if not target_rows_by_key:
            return None

        matched_rows: list[dict[str, object]] = []
        for source_row_index, source_row in enumerate(source_table.rows, start=1):
            key_value = source_row.get(source_key_header)
            normalized_key = self._normalize_key(key_value)
            target_row_index = target_rows_by_key.get(normalized_key)
            if target_row_index is None:
                continue
            matched_rows.append(
                {
                    "sourceRowIndex": source_row_index,
                    "sourceRow": source_row,
                    "targetRowIndex": target_row_index,
                    "keyValue": key_value,
                }
            )

        if not matched_rows:
            return None

        fill_mappings = self._select_keyed_fill_mappings(
            target_sheet=target_sheet,
            matched_rows=matched_rows,
            column_mappings=column_mappings,
            key_mapping=key_mapping,
        )
        if not fill_mappings:
            return None

        return {
            "keyMapping": key_mapping,
            "matchedRows": matched_rows,
            "fillMappings": fill_mappings,
        }

    def _select_key_mapping(self, column_mappings: list[dict[str, object]]) -> dict[str, object] | None:
        candidates: list[tuple[int, dict[str, object]]] = []
        for mapping in column_mappings:
            source_header = mapping.get("sourceHeader")
            target_header = mapping.get("targetHeader")
            source_priority = self._key_header_priority(source_header)
            target_priority = self._key_header_priority(target_header)
            if source_priority is None or target_priority is None:
                continue
            candidates.append((source_priority + target_priority, mapping))

        if not candidates:
            return None
        candidates.sort(key=lambda item: (item[0], str(item[1].get("targetHeader") or "")))
        return candidates[0][1]

    def _index_target_rows_by_key(
        self,
        *,
        target_sheet: TargetSheet,
        target_key_column: int,
    ) -> dict[str, int]:
        rows_by_key: dict[str, int] = {}
        for row_index in range(target_sheet.header_row_index + 1, target_sheet.worksheet.max_row + 1):
            key = self._normalize_key(
                target_sheet.worksheet.cell(row=row_index, column=target_key_column).value
            )
            if key and key not in rows_by_key:
                rows_by_key[key] = row_index
        return rows_by_key

    def _select_keyed_fill_mappings(
        self,
        *,
        target_sheet: TargetSheet,
        matched_rows: list[dict[str, object]],
        column_mappings: list[dict[str, object]],
        key_mapping: dict[str, object],
    ) -> list[dict[str, object]]:
        key_target_column = int(key_mapping["targetColumnIndex"])
        fill_mappings: list[dict[str, object]] = []
        for mapping in column_mappings:
            target_column = int(mapping["targetColumnIndex"])
            if target_column == key_target_column:
                continue
            if not self._has_fillable_matched_cell(
                target_sheet=target_sheet,
                matched_rows=matched_rows,
                mapping=mapping,
            ):
                continue
            fill_mappings.append(mapping)
        return fill_mappings

    def _has_fillable_matched_cell(
        self,
        *,
        target_sheet: TargetSheet,
        matched_rows: list[dict[str, object]],
        mapping: dict[str, object],
    ) -> bool:
        source_header = str(mapping["sourceHeader"])
        target_column = int(mapping["targetColumnIndex"])
        for matched_row in matched_rows:
            source_row = matched_row.get("sourceRow")
            if not isinstance(source_row, dict):
                continue
            if source_row.get(source_header) in (None, ""):
                continue
            target_row_index = int(matched_row["targetRowIndex"])
            if self._is_empty_like(
                target_sheet.worksheet.cell(row=target_row_index, column=target_column).value
            ):
                return True
        return False

    def _write_keyed_rows(
        self,
        *,
        source_table: SourceTable,
        target_sheet: TargetSheet,
        fill_plan: dict[str, object],
        mode: str,
    ) -> list[dict[str, object]]:
        audit: list[dict[str, object]] = []
        key_mapping = dict(fill_plan["keyMapping"])
        matched_rows = [
            item for item in fill_plan.get("matchedRows", []) or []
            if isinstance(item, dict) and isinstance(item.get("sourceRow"), dict)
        ]
        fill_mappings = [
            item for item in fill_plan.get("fillMappings", []) or []
            if isinstance(item, dict)
        ]
        key_source_header = str(key_mapping.get("sourceHeader") or "")

        for matched_row in matched_rows:
            source_row = dict(matched_row["sourceRow"])
            source_row_index = int(matched_row["sourceRowIndex"])
            target_row_index = int(matched_row["targetRowIndex"])
            key_value = matched_row.get("keyValue")

            for mapping in fill_mappings:
                source_header = str(mapping["sourceHeader"])
                column_index = int(mapping["targetColumnIndex"])
                value = source_row.get(source_header)
                cell = target_sheet.worksheet.cell(row=target_row_index, column=column_index)
                existing_value = cell.value
                status = "source_empty"
                if value not in (None, ""):
                    if self._is_empty_like(existing_value):
                        cell.value = value
                        status = "written"
                    elif self._cell_values_equal(existing_value, value):
                        status = "kept_existing"
                    else:
                        status = "preserved_existing"

                audit.append(
                    {
                        "sheet": target_sheet.worksheet.title,
                        "cell": cell.coordinate,
                        "status": status,
                        "mode": mode,
                        "sourceDocument": getattr(source_table.document, "name", ""),
                        "sourceSheet": source_table.sheet.get("title"),
                        "sourceRow": source_row_index,
                        "sourceColumn": source_header,
                        "targetColumn": mapping.get("targetHeader"),
                        "targetRow": target_row_index,
                        "matchKeySourceColumn": key_source_header,
                        "matchKeyValue": key_value,
                        "value": value,
                        "existingValue": existing_value,
                        "decisionSource": "auto",
                        "writePolicy": "only_empty",
                        "candidateScore": mapping.get("score"),
                        "riskLevel": "low" if float(mapping.get("score", 0.0) or 0.0) >= 1.0 else "medium",
                        "reasonSummary": (
                            f"按关键字段“{key_source_header}”匹配到目标已有行；"
                            f"{mapping.get('reason') or ''}"
                        ).strip("；"),
                        "message": self._build_audit_message(
                            status=status,
                            sheet=target_sheet.worksheet.title,
                            cell=cell.coordinate,
                            source_header=source_header,
                        ),
                    }
                )

        written_rows = {
            int(item["targetRow"])
            for item in audit
            if str(item.get("status") or "") == "written" and item.get("targetRow")
        }
        for row_index in written_rows:
            target_sheet.worksheet.row_dimensions[row_index].hidden = False
        return audit

    def _write_rows(
        self,
        *,
        source_table: SourceTable,
        target_sheet: TargetSheet,
        column_mappings: list[dict[str, object]],
        mode: str,
    ) -> list[dict[str, object]]:
        audit: list[dict[str, object]] = []
        target_row = self._find_first_available_row(target_sheet, column_mappings)
        rows_written: set[int] = set()

        for source_row_index, source_row in enumerate(source_table.rows, start=1):
            write_row = self._find_next_available_row(
                target_sheet=target_sheet,
                column_mappings=column_mappings,
                start_row=target_row,
            )
            target_row = write_row + 1
            row_has_written_value = False

            for mapping in column_mappings:
                source_header = str(mapping["sourceHeader"])
                column_index = int(mapping["targetColumnIndex"])
                value = source_row.get(source_header)
                cell = target_sheet.worksheet.cell(row=write_row, column=column_index)
                existing_value = cell.value
                status = "source_empty"
                if value not in (None, ""):
                    if self._is_empty_like(existing_value):
                        cell.value = value
                        status = "written"
                        row_has_written_value = True
                    elif self._cell_values_equal(existing_value, value):
                        status = "kept_existing"
                    else:
                        status = "preserved_existing"

                audit.append(
                    {
                        "sheet": target_sheet.worksheet.title,
                        "cell": cell.coordinate,
                        "status": status,
                        "mode": mode,
                        "sourceDocument": getattr(source_table.document, "name", ""),
                        "sourceSheet": source_table.sheet.get("title"),
                        "sourceRow": source_row_index,
                        "sourceColumn": source_header,
                        "targetColumn": mapping.get("targetHeader"),
                        "value": value,
                        "existingValue": existing_value,
                        "decisionSource": "auto",
                        "writePolicy": "only_empty",
                        "candidateScore": mapping.get("score"),
                        "riskLevel": "low" if float(mapping.get("score", 0.0) or 0.0) >= 1.0 else "medium",
                        "reasonSummary": mapping.get("reason"),
                        "message": self._build_audit_message(
                            status=status,
                            sheet=target_sheet.worksheet.title,
                            cell=cell.coordinate,
                            source_header=source_header,
                        ),
                    }
                )

            if row_has_written_value:
                rows_written.add(write_row)

        for row_index in rows_written:
            target_sheet.worksheet.row_dimensions[row_index].hidden = False
        return audit

    def _find_first_available_row(
        self,
        target_sheet: TargetSheet,
        column_mappings: list[dict[str, object]],
    ) -> int:
        return max(target_sheet.header_row_index + 1, 1)

    def _find_next_available_row(
        self,
        *,
        target_sheet: TargetSheet,
        column_mappings: list[dict[str, object]],
        start_row: int,
    ) -> int:
        row_index = start_row
        while row_index <= max(target_sheet.worksheet.max_row, target_sheet.header_row_index) + 1:
            if all(
                self._is_empty_like(
                    target_sheet.worksheet.cell(
                        row=row_index,
                        column=int(mapping["targetColumnIndex"]),
                    ).value
                )
                for mapping in column_mappings
            ):
                return row_index
            row_index += 1
        return row_index

    def _summarize_audit(self, audit: list[dict[str, object]], *, mode: str) -> dict[str, object]:
        stats = {
            "mode": mode,
            "written": 0,
            "keptExisting": 0,
            "preservedExisting": 0,
            "sourceEmpty": 0,
            "rowsTransferred": 0,
        }
        written_rows: set[int] = set()
        for item in audit:
            status = str(item.get("status") or "")
            if status == "written":
                stats["written"] += 1
                cell = str(item.get("cell") or "")
                row_digits = "".join(character for character in cell if character.isdigit())
                if row_digits:
                    written_rows.add(int(row_digits))
            elif status == "kept_existing":
                stats["keptExisting"] += 1
            elif status == "preserved_existing":
                stats["preservedExisting"] += 1
            elif status == "source_empty":
                stats["sourceEmpty"] += 1
        stats["rowsTransferred"] = len(written_rows)
        return stats

    def _build_audit_message(self, *, status: str, sheet: str, cell: str, source_header: str) -> str:
        if status == "written":
            return f"已把源列“{source_header}”写入工作表“{sheet}”的 {cell}。"
        if status == "kept_existing":
            return f"工作表“{sheet}”的 {cell} 已有相同值，无需重复写入。"
        if status == "preserved_existing":
            return f"工作表“{sheet}”的 {cell} 已有不同值，已保留原值。"
        return f"源列“{source_header}”为空，未写入工作表“{sheet}”的 {cell}。"

    def _encode_workbook(self, workbook: Workbook) -> str:
        buffer = BytesIO()
        workbook.save(buffer)
        return base64.b64encode(buffer.getvalue()).decode("utf-8")

    def _build_export_filename(self, target_documents: list[Any]) -> str:
        if target_documents:
            filename = str(getattr(target_documents[0], "name", "") or "").strip()
            if filename:
                stem, _extension = os.path.splitext(filename)
                return f"{stem or '目标表格'}_filled.xlsx"
        return "workflow_transfer_filled.xlsx"

    def _make_unique_sheet_name(self, workbook: Workbook, raw_title: str) -> str:
        base = str(raw_title or "Sheet").strip()[:31] or "Sheet"
        candidate = base
        suffix = 2
        while candidate in workbook.sheetnames:
            suffix_text = f"_{suffix}"
            candidate = f"{base[: 31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        return candidate

    def _key_header_priority(self, value: object) -> int | None:
        normalized = self._normalize_header(value)
        if not normalized:
            return None

        exact_groups = [
            {"id", "key", "code", "编号", "编码", "代码", "标识", "唯一标识", "指标编码", "项目编码"},
            {"单号", "订单号", "合同号", "发票号", "凭证号", "供应商编码", "客户编码"},
            {"名称", "name", "项目", "指标", "科目", "esg指标", "字段", "条目"},
            {"序号", "行号", "no", "number", "index"},
        ]
        for priority, values in enumerate(exact_groups):
            if normalized in values:
                return priority

        contains_groups = [
            ("唯一", "编码", "编号", "代码", "id", "code"),
            ("单号", "订单号", "合同号", "发票号", "凭证号"),
            ("名称", "name", "项目", "指标", "科目", "字段", "条目"),
            ("序号", "行号"),
        ]
        for priority, fragments in enumerate(contains_groups):
            if any(fragment in normalized for fragment in fragments):
                return priority + 4
        return None

    def _normalize_header(self, value: object) -> str:
        return "".join(
            character
            for character in str(value or "").strip().lower()
            if character.isalnum() or "\u4e00" <= character <= "\u9fff"
        )

    def _normalize_key(self, value: object) -> str:
        return self._normalize_header(value)

    def _is_empty_like(self, value: object) -> bool:
        if value is None:
            return True
        return self._normalize_header(value) in {self._normalize_header(item) for item in PLACEHOLDER_CELL_VALUES}

    def _cell_values_equal(self, left: object, right: object) -> bool:
        if left == right:
            return True
        return self._normalize_header(left) == self._normalize_header(right)
