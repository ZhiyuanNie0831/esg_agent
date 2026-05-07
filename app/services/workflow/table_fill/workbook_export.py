"""Excel workbook 导出工具。

负责把已经整理好的结果行写回工作簿，并生成可下载的 `.xlsx` 内容。
主要职责有三件事：

1. 决定往哪个 workbook 里写
   - 优先使用原始上传的工作簿，尽量保留模板结构
   - 拿不到原始文件时，再根据结构化数据重建 workbook
2. 决定写到哪个 sheet
   - 优先复用像 `指标 / 结果` 这样的模板 sheet
   - 没有合适模板时，退回到系统生成的 `结果表`
3. 把最终 workbook 序列化成 API 可返回的下载内容
"""

from __future__ import annotations

import base64
import logging
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.workflow.table_fill.crosscheck import TableFillMappingCrossCheckService
from app.services.workflow.uploads.store import workflow_upload_store

RESULT_SHEET_TITLE = "结果表"
XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PLACEHOLDER_CELL_VALUES = {"", "-", "--", "待填", "待填写", "待补充", "tbd", "n/a", "na", "pending"}
LABEL_SCAN_ROW_BASE_LIMIT = 40
LABEL_SCAN_ROW_MAX_LIMIT = 1000
LABEL_SCAN_COLUMN_BASE_LIMIT = 8
LABEL_SCAN_COLUMN_MAX_LIMIT = 50
logger = logging.getLogger(__name__)
TEMPLATE_DOCUMENT_HINTS = (
    "template",
    "summary",
    "result",
    "report",
    "form",
    "blank",
    "模板",
    "汇总",
    "填报",
    "填表",
    "回填",
    "结果",
    "空白",
    "目标",
    "目标表",
    "目标表格",
)
SOURCE_DOCUMENT_HINTS = (
    "detail",
    "data",
    "source",
    "ledger",
    "raw",
    "明细",
    "台账",
    "原始",
    "数据",
    "数据源",
    "源表",
)
TEMPLATE_SHEET_HINTS = ("summary", "result", "report", "form", "汇总", "模板", "填报", "结果")
METRIC_LABEL_ALIASES = {
    "sum": {"sum", "total", "合计", "总计", "汇总", "求和", "总和", "总额", "总量", "小计", "累计"},
    "avg": {"avg", "average", "mean", "平均", "平均值"},
    "max": {"max", "maximum", "最大", "最高"},
    "min": {"min", "minimum", "最小", "最低"},
    "count": {"count", "数量", "条数", "行数", "计数"},
    "ratio": {"ratio", "占比", "比例"},
    "intensity": {"intensity", "强度", "密度"},
    "yoy": {"yoy", "同比", "yearoveryear"},
    "mom": {"mom", "环比", "monthovermonth"},
}
CONTEXT_LABEL_ALIASES = {
    "amount": {"金额", "费用", "成本", "支出", "amount", "cost", "expense"},
    "emission": {"排放", "排放量", "碳排放", "emission", "emissions"},
    "energy": {"能源", "能耗", "用电", "电量", "energy", "electricity"},
    "water": {"水", "用水", "耗水", "取水", "water"},
    "waste": {"废弃物", "废物", "垃圾", "waste"},
    "revenue": {"收入", "营收", "营业收入", "revenue"},
    "production": {"产量", "产出", "生产量", "production", "output"},
    "employee": {"员工", "人数", "雇员", "headcount", "employee"},
}
TITLE_LABEL_HINTS = ("汇总", "模板", "报表", "统计表", "数据表", "披露", "说明")
HEADER_ALIASES = {
    "指标": {"指标", "metric", "metrics"},
    "结果": {"结果", "value", "result", "数值"},
    "文档": {"文档", "document", "file"},
    "工作表": {"工作表", "sheet", "worksheet"},
    "数值字段": {"数值字段", "字段", "column", "metric field"},
    "分组字段": {"分组字段", "group by", "group_field"},
    "分组值": {"分组值", "group value", "group"},
}


@dataclass(frozen=True)
class HeaderLayout:
    """工作表表头布局信息。

    - `row_index`：哪一行被识别为表头
    - `headers`：表头原始顺序
    - `positions`：表头名称到 Excel 列号的映射
    """

    row_index: int
    headers: list[str]
    positions: dict[str, int]


@dataclass(frozen=True)
class TemplateFillReport:
    """描述一次模板回填的结果。"""

    applied: bool
    mode: str
    audit: list[dict[str, object]]


class TableFillWorkbookExporter:
    """根据结果行生成可下载的 Excel 工作簿。"""

    def __init__(self, crosscheck_service: TableFillMappingCrossCheckService | None = None) -> None:
        self._crosscheck_service = crosscheck_service or TableFillMappingCrossCheckService()

    def build_export_bundle(
        self,
        *,
        task: str,
        documents: list[Any],
        headers: list[str],
        rows: list[dict[str, object]],
        manual_mappings: list[dict[str, object]] | None = None,
        crosscheck_result: dict[str, object] | None = None,
    ) -> dict[str, object]:
        """生成导出文件以及回填审计结果。"""
        fill_plan = self._prepare_fill_plan(
            task=task,
            documents=documents,
            rows=rows,
        )
        workbook_documents = list(fill_plan["workbookDocuments"])
        workbook = fill_plan["workbook"]
        target_sheet = fill_plan.get("targetWorksheet")
        used_original_workbook = bool(fill_plan["usedOriginalWorkbook"])
        mapping_candidates = list(fill_plan.get("mappingCandidates", []))
        if manual_mappings:
            mapping_candidates = self._apply_manual_targets_to_review_candidates(
                mapping_candidates=mapping_candidates,
                manual_mappings=manual_mappings,
            )
        crosscheck = crosscheck_result or self._review_mapping_candidates(
            task=task,
            fill_plan=fill_plan,
            mapping_candidates=mapping_candidates,
        )
        mapping_candidates = self._apply_crosscheck_to_mapping_candidates(
            mapping_candidates=mapping_candidates,
            crosscheck=crosscheck,
        )

        if bool(crosscheck.get("blockWrite")):
            audit = [
                self._build_crosscheck_blocked_audit(
                    fill_plan=fill_plan,
                    crosscheck=crosscheck,
                )
            ]
            return {
                "exportFiles": [],
                "fillAudit": audit,
                "fillStats": self._summarize_fill_audit(
                    TemplateFillReport(applied=False, mode="blocked_by_crosscheck", audit=audit)
                ),
                "crossCheck": crosscheck,
            }

        fill_report = TemplateFillReport(applied=False, mode="unresolved_template", audit=[])
        handled_mapping_ids: set[str] = set()
        manual_mapping_ids = {
            str(item.get("mappingId") or "").strip()
            for item in manual_mappings or []
            if isinstance(item, dict) and str(item.get("mappingId") or "").strip()
        }
        unresolved_required_candidates = [
            item
            for item in mapping_candidates
            if bool(item.get("requiresConfirmation")) and str(item.get("mappingId") or "").strip() not in manual_mapping_ids
        ]

        if unresolved_required_candidates:
            fill_report = TemplateFillReport(
                applied=False,
                mode="requires_confirmation",
                audit=[
                    self._build_unconfirmed_candidate_audit(item)
                    for item in unresolved_required_candidates
                ],
            )

        if manual_mappings:
            manual_report, handled_mapping_ids = self._apply_manual_mappings(
                workbook=workbook,
                rows=rows,
                manual_mappings=manual_mappings,
                mapping_candidates=mapping_candidates,
            )
            fill_report = self._merge_fill_reports(fill_report, manual_report)

        if target_sheet is not None and not unresolved_required_candidates:
            auto_report = self._apply_mapping_candidates(
                workbook=workbook,
                rows=rows,
                mapping_candidates=mapping_candidates,
                handled_mapping_ids=handled_mapping_ids,
            )
            fill_report = self._merge_fill_reports(fill_report, auto_report)

        if target_sheet is not None and fill_report.applied:
            export_summary = self._build_template_fill_summary(
                sheet_title=target_sheet.title,
                used_original_workbook=used_original_workbook,
            )
        else:
            result_sheet = self._replace_result_sheet(workbook)
            self._write_result_sheet(result_sheet, headers, rows)
            export_summary = self._build_result_sheet_summary(
                sheet_title=result_sheet.title,
                used_original_workbook=used_original_workbook,
            )
            fill_report = TemplateFillReport(
                applied=True,
                mode="result_sheet",
                audit=[
                    {
                        "sheet": result_sheet.title,
                        "cell": None,
                        "status": "result_sheet_created",
                        "field": None,
                        "metric": None,
                        "value": None,
                        "existingValue": None,
                        "sourceDocument": None,
                        "sourceSheet": None,
                        "sourceColumn": None,
                        "decisionSource": "system",
                        "writePolicy": "only_empty",
                        "candidateScore": None,
                        "riskLevel": "medium" if unresolved_required_candidates else "low",
                        "reasonSummary": self._build_result_sheet_reason(fill_report.audit),
                        "message": f"模板未能稳定定位，结果已写入工作表“{result_sheet.title}”。",
                    }
                ],
            )

        export_files = [
            {
                "label": export_summary,
                "filename": self._build_export_filename(workbook_documents),
                "mimeType": XLSX_MIME_TYPE,
                "contentBase64": self._encode_workbook(workbook),
            }
        ]
        return {
            "exportFiles": export_files,
            "fillAudit": fill_report.audit,
            "fillStats": self._summarize_fill_audit(fill_report),
            "crossCheck": crosscheck,
        }

    def preview_fill_plan(
        self,
        *,
        task: str,
        documents: list[Any],
        rows: list[dict[str, object]],
    ) -> dict[str, object]:
        """生成候选填位预览，供人工确认前编辑。"""
        fill_plan = self._prepare_fill_plan(task=task, documents=documents, rows=rows)
        mapping_candidates = list(fill_plan.get("mappingCandidates", []))
        crosscheck = self._review_mapping_candidates(
            task=task,
            fill_plan=fill_plan,
            mapping_candidates=mapping_candidates,
        )
        mapping_candidates = self._apply_crosscheck_to_mapping_candidates(
            mapping_candidates=mapping_candidates,
            crosscheck=crosscheck,
        )
        return {
            "mode": fill_plan["mode"],
            "targetWorkbook": self._build_export_filename(list(fill_plan["workbookDocuments"])),
            "targetSheet": fill_plan["targetSheet"],
            "mappingCandidates": mapping_candidates,
            "crossCheck": crosscheck,
            "summary": str(fill_plan.get("summary") or "").strip(),
        }

    def build_export_files(
        self,
        *,
        task: str,
        documents: list[Any],
        headers: list[str],
        rows: list[dict[str, object]],
    ) -> list[dict[str, str]]:
        """兼容旧调用方，只返回导出文件列表。"""
        return list(
            self.build_export_bundle(
                task=task,
                documents=documents,
                headers=headers,
                rows=rows,
            ).get("exportFiles", [])
        )

    def _prepare_fill_plan(
        self,
        *,
        task: str,
        documents: list[Any],
        rows: list[dict[str, object]],
    ) -> dict[str, object]:
        source_documents = self._select_source_documents(documents, rows)
        target_documents = self._select_target_documents(
            task=task,
            documents=documents,
            rows=rows,
        )
        workbook_documents = target_documents or source_documents or documents
        workbook, source_sheet_names, used_original_workbook = self._build_workbook_base(workbook_documents)
        target_sheet = self._choose_result_sheet(
            workbook=workbook,
            task=task,
            source_sheet_names=source_sheet_names,
            rows=rows,
        )
        mapping_rows = self._build_mapping_rows(rows)

        if target_sheet is None:
            candidates = [
                self._build_mapping_candidate(
                    mapping_id=mapping_id,
                    row=row,
                    sheet=RESULT_SHEET_TITLE,
                    cell="",
                    status="result_sheet_fallback",
                    mode="result_sheet",
                    confidence="low",
                    score=0.18,
                    risk_level="high",
                    requires_confirmation=True,
                    reasons=["未识别到稳定模板，自动回填风险较高。"],
                    alternative_candidates=[],
                )
                for mapping_id, row in mapping_rows
            ]
            return {
                "mode": "result_sheet",
                "workbookDocuments": workbook_documents,
                "workbook": workbook,
                "usedOriginalWorkbook": used_original_workbook,
                "targetWorksheet": None,
                "targetSheet": RESULT_SHEET_TITLE,
                "mappingCandidates": candidates,
                "summary": "未识别到稳定模板，默认会回退为新增结果表；你也可以在确认前手工指定单元格。",
            }

        header_layout = self._detect_header_layout(target_sheet)
        if header_layout is not None and header_layout.positions.get("结果"):
            candidates = self._preview_structured_template_sheet(
                worksheet=target_sheet,
                header_layout=header_layout,
                rows=rows,
            )
            mode = "header_layout"
        else:
            candidates = self._preview_label_value_template_sheet(
                worksheet=target_sheet,
                rows=rows,
            )
            mode = "label_value"

        matched_count = len([item for item in candidates if str(item.get("cell") or "").strip()])
        confirmation_count = len([item for item in candidates if bool(item.get("requiresConfirmation"))])
        summary = f"已为 {matched_count} 条结果生成候选填位，可在确认前人工修改。"
        if confirmation_count:
            summary += f" 其中有 {confirmation_count} 条低置信度候选，建议逐项确认或跳过。"
        return {
            "mode": mode,
            "workbookDocuments": workbook_documents,
            "workbook": workbook,
            "usedOriginalWorkbook": used_original_workbook,
            "targetWorksheet": target_sheet,
            "targetSheet": target_sheet.title,
            "mappingCandidates": candidates,
            "summary": summary,
        }

    # workbook 基础准备

    def _build_workbook_base(
        self,
        documents: list[Any],
    ) -> tuple[Workbook, list[str], bool]:
        """准备后续回填要写入的 workbook。

        返回值依次表示：
        - workbook：要写入的工作簿对象
        - source_sheet_names：来自输入材料的 sheet 名称
        - bool：当前是否保留了原始上传 workbook
        """
        original_workbook = self._load_original_workbook(documents)
        if original_workbook is not None:
            return original_workbook, list(original_workbook.sheetnames), True

        # 如果没有原始 workbook，就根据结构化 sheet 数据重建一份。
        # 这条路径常见于手工构造请求或上传字节已失效的情况。
        workbook = Workbook()
        workbook.remove(workbook.active)
        source_sheet_names: list[str] = []

        for document in documents:
            if document.type != "excel":
                continue

            for sheet in document.structuredData.get("sheets", []):
                worksheet = workbook.create_sheet(
                    title=self._make_unique_sheet_name(
                        workbook,
                        str(sheet.get("title", "Sheet")),
                    )
                )
                source_sheet_names.append(worksheet.title)
                self._write_source_sheet(worksheet, sheet)

        if not workbook.worksheets:
            workbook.create_sheet(title=RESULT_SHEET_TITLE)

        return workbook, source_sheet_names, False

    def _load_original_workbook(self, documents: list[Any]) -> Workbook | None:
        """优先取回最可能的原始上传 workbook。

        即使本次请求里有多份 Excel，只要结果明确来自其中一份，也尽量保留那份原始模板。
        """
        excel_documents = [document for document in documents if document.type == "excel"]
        if len(excel_documents) != 1:
            return None

        token = str(excel_documents[0].structuredData.get("workbookUploadToken") or "").strip()
        workbook_bytes = workflow_upload_store.get_bytes(token)
        if not workbook_bytes:
            return None

        try:
            return load_workbook(BytesIO(workbook_bytes))
        except Exception:
            logger.exception(
                "table fill original workbook load failed; falling back to structured workbook: document=%s",
                getattr(excel_documents[0], "name", ""),
            )
            return None

    def _select_source_documents(
        self,
        documents: list[Any],
        rows: list[dict[str, object]],
    ) -> list[Any]:
        """收窄用于回填的 Excel 文档集合，避免多文件时把无关工作簿也混进来。"""
        excel_documents = [document for document in documents if document.type == "excel"]
        if not excel_documents:
            return []

        preferred_name_set = {
            str(row.get("文档") or "").strip()
            for row in rows
            if str(row.get("文档") or "").strip()
        }
        matched_documents = [
            document
            for document in excel_documents
            if str(getattr(document, "name", "") or "").strip() in preferred_name_set
        ]
        return matched_documents or excel_documents

    def _select_target_documents(
        self,
        *,
        task: str,
        documents: list[Any],
        rows: list[dict[str, object]],
    ) -> list[Any]:
        """优先选择更像“回填模板”的 workbook，而不是默认回写到源数据文件。"""
        excel_documents = [document for document in documents if document.type == "excel"]
        if len(excel_documents) <= 1:
            return excel_documents

        task_lower = str(task or "").lower()
        source_name_set = {
            str(row.get("文档") or "").strip()
            for row in rows
            if str(row.get("文档") or "").strip()
        }
        scored_documents = sorted(
            (
                self._score_target_document(
                    document,
                    task_lower=task_lower,
                    source_name_set=source_name_set,
                ),
                index,
                document,
            )
            for index, document in enumerate(excel_documents)
        )
        best_score = scored_documents[-1][0]
        if best_score <= 0:
            non_source_documents = [
                document
                for document in excel_documents
                if str(getattr(document, "name", "") or "").strip() not in source_name_set
            ]
            if len(non_source_documents) == 1:
                return non_source_documents
            return self._select_source_documents(documents, rows)[:1]

        return [scored_documents[-1][2]]

    def _score_target_document(
        self,
        document: Any,
        *,
        task_lower: str,
        source_name_set: set[str],
    ) -> int:
        """对候选 workbook 打分，分数越高越像应被回填的模板文件。"""
        name = str(getattr(document, "name", "") or "").strip()
        lowered_name = name.lower()
        stem = lowered_name.rsplit(".", 1)[0] if "." in lowered_name else lowered_name
        score = 0

        if lowered_name and lowered_name in task_lower:
            score += 8
        elif stem and stem in task_lower:
            score += 6

        score += 3 * sum(1 for hint in TEMPLATE_DOCUMENT_HINTS if hint in lowered_name)
        score -= 2 * sum(1 for hint in SOURCE_DOCUMENT_HINTS if hint in lowered_name)
        score += self._score_document_template_structure(document)

        if name and name in source_name_set:
            score -= 2
        elif name:
            score += 2

        return score

    def _score_document_template_structure(self, document: Any) -> int:
        """根据结构化 sheet 判断文档是否像回填模板。"""
        best_score = 0
        for sheet in document.structuredData.get("sheets", []):
            if not isinstance(sheet, dict):
                continue
            best_score = max(best_score, self._score_structured_template_sheet(sheet))
        return best_score

    def _score_structured_template_sheet(self, sheet: dict[str, Any]) -> int:
        """对单个结构化 sheet 打分，识别“指标/结果”类模板布局。"""
        sheet_name = str(sheet.get("title", "") or "")
        headers = [str(header) for header in sheet.get("headers", []) if str(header).strip()]
        metric_header = self._resolve_structured_header(headers, canonical="指标")
        result_header = self._resolve_structured_header(headers, canonical="结果")
        label_anchor_count = len(sheet.get("labelAnchors", []) or [])
        empty_value_zone_count = len(sheet.get("emptyValueZones", []) or [])
        sheet_role = str(sheet.get("sheetRole") or "").strip()
        header_confidence = str(sheet.get("headerConfidence") or "").strip()

        score = 0
        if any(hint in sheet_name.lower() for hint in TEMPLATE_SHEET_HINTS):
            score += 2
        if metric_header and result_header:
            score += 5
        elif result_header:
            score += 3
        elif metric_header:
            score += 2
        if sheet_role == "template_like":
            score += 3
        elif sheet_role == "source_like":
            score -= 2
        if header_confidence == "high":
            score += 1
        elif header_confidence == "low":
            score -= 1
        score += min(label_anchor_count, 3) * 2
        score += min(empty_value_zone_count, 3)
        if any(
            self._matches_header_alias(header, canonical)
            for header in headers
            for canonical in ("数值字段", "工作表", "文档")
        ):
            score += 1

        if metric_header and result_header:
            rows = sheet.get("rows", [])
            if any(
                isinstance(row, dict)
                and row.get(metric_header) not in (None, "")
                and row.get(result_header) in (None, "")
                for row in rows
            ):
                score += 2

        return score

    def _resolve_structured_header(self, headers: list[str], *, canonical: str) -> str | None:
        """在结构化表头中找到某个规范字段对应的原始列名。"""
        return next(
            (
                header
                for header in headers
                if self._matches_header_alias(header, canonical)
            ),
            None,
        )

    def _matches_header_alias(self, header: str, canonical: str) -> bool:
        """判断表头是否命中某个规范字段的别名集合。"""
        normalized = str(header or "").strip().lower()
        if not normalized:
            return False
        aliases = {canonical.lower(), *(alias.lower() for alias in HEADER_ALIASES.get(canonical, set()))}
        return normalized in aliases

    def _write_source_sheet(self, worksheet: Worksheet, structure: dict[str, Any]) -> None:
        """把结构化 sheet 内容写入重建后的 worksheet。"""
        headers = [str(header) for header in structure.get("headers", []) if str(header).strip()]
        if not headers:
            for merged_range in structure.get("mergedRanges", []) or []:
                range_ref = str((merged_range or {}).get("range") if isinstance(merged_range, dict) else merged_range).strip()
                if range_ref:
                    worksheet.merge_cells(range_ref)
            return

        worksheet.append(headers)
        for row in structure.get("rows", []):
            if not isinstance(row, dict):
                continue
            worksheet.append([row.get(header) for header in headers])
        for merged_range in structure.get("mergedRanges", []) or []:
            range_ref = str((merged_range or {}).get("range") if isinstance(merged_range, dict) else merged_range).strip()
            if range_ref:
                worksheet.merge_cells(range_ref)

    def _encode_workbook(self, workbook: Workbook) -> str:
        """把 workbook 编码成 base64，便于通过 JSON 返回。"""
        buffer = BytesIO()
        workbook.save(buffer)
        return base64.b64encode(buffer.getvalue()).decode("utf-8")

    # 结果 sheet 选择与回填

    def _choose_result_sheet(
        self,
        *,
        workbook: Workbook,
        task: str,
        source_sheet_names: list[str],
        rows: list[dict[str, object]],
    ) -> Worksheet | None:
        """挑选最适合写入结果的 worksheet。

        这里只会评估来自输入材料的 sheet，目标是尽量复用用户已有模板，
        而不是每次都新建 `结果表`。
        """
        task_lower = str(task or "").lower()
        candidate_sheets: list[tuple[int, Worksheet]] = []

        for sheet_name in source_sheet_names:
            worksheet = workbook[sheet_name]
            header_layout = self._detect_header_layout(worksheet)
            if header_layout is None:
                continue

            score = self._score_template_sheet(
                worksheet=worksheet,
                sheet_name=sheet_name,
                headers=header_layout.headers,
                task_lower=task_lower,
                rows=rows,
            )
            if score > 0:
                candidate_sheets.append((score, worksheet))

        if not candidate_sheets:
            return None

        candidate_sheets.sort(key=lambda item: (-item[0], item[1].title))
        return candidate_sheets[0][1]

    def _score_template_sheet(
        self,
        *,
        worksheet: Worksheet,
        sheet_name: str,
        headers: list[str],
        task_lower: str,
        rows: list[dict[str, object]],
    ) -> int:
        """为候选 worksheet 打分，判断它是否像结果模板。"""
        has_metric = any(self._matches_header_alias(header, "指标") for header in headers)
        has_result = any(self._matches_header_alias(header, "结果") for header in headers)
        score = 0
        if has_result and has_metric:
            score += 5
        elif has_result:
            score += 3
        elif has_metric:
            score += 2
        if any(
            self._matches_header_alias(header, canonical)
            for header in headers
            for canonical in ("数值字段", "工作表", "文档")
        ):
            score += 1
        sheet_name_lower = sheet_name.lower()
        if any(hint in sheet_name_lower for hint in TEMPLATE_SHEET_HINTS):
            score += 2
        if sheet_name_lower and sheet_name_lower in task_lower:
            score += 2
        score += self._score_metric_label_matches(worksheet, rows)
        return score

    def _fill_existing_template_sheet(
        self,
        worksheet: Worksheet,
        rows: list[dict[str, object]],
    ) -> TemplateFillReport:
        """把结果行回填到一个现有模板 worksheet 中。

        如果模板里已经有 `sum`、`avg` 这类指标行，会优先按指标名复用原行，
        从而尽量不破坏原有版式。
        """
        header_layout = self._detect_header_layout(worksheet)
        if header_layout is not None and header_layout.positions.get("结果"):
            return self._fill_structured_template_sheet(
                worksheet=worksheet,
                header_layout=header_layout,
                rows=rows,
            )

        return self._fill_label_value_template_sheet(
            worksheet=worksheet,
            rows=rows,
        )

    def _apply_manual_mappings(
        self,
        *,
        workbook: Workbook,
        rows: list[dict[str, object]],
        manual_mappings: list[dict[str, object]],
        mapping_candidates: list[dict[str, object]],
    ) -> tuple[TemplateFillReport, set[str]]:
        """优先应用人工确认后的单元格映射。"""
        candidate_index = {
            str(item.get("mappingId") or "").strip(): item
            for item in mapping_candidates
            if isinstance(item, dict) and str(item.get("mappingId") or "").strip()
        }
        mappings_by_id = {
            str(item.get("mappingId") or "").strip(): item
            for item in manual_mappings
            if isinstance(item, dict) and str(item.get("mappingId") or "").strip()
        }
        if not mappings_by_id:
            return TemplateFillReport(applied=False, mode="manual_mapping", audit=[]), set()

        audit: list[dict[str, object]] = []
        handled_mapping_ids: set[str] = set()
        applied = False

        for mapping_id, row in self._build_mapping_rows(rows):
            manual_item = mappings_by_id.get(mapping_id)
            if manual_item is None:
                continue
            handled_mapping_ids.add(mapping_id)

            if manual_item.get("enabled") is False:
                audit.append(
                    self._build_skipped_manual_mapping_audit(
                        mapping_id=mapping_id,
                        row=row,
                    )
                )
                continue

            write_policy = self._normalize_write_policy(manual_item.get("writePolicy"))
            candidate_item = candidate_index.get(mapping_id)
            selected_option = self._resolve_manual_target_option(
                manual_item=manual_item,
                candidate_item=candidate_item,
            )
            sheet_name = str(selected_option.get("sheet") or manual_item.get("sheet") or "").strip()
            cell_ref = str(selected_option.get("cell") or manual_item.get("cell") or "").strip().upper()
            if not sheet_name or not cell_ref or sheet_name not in workbook.sheetnames or not self._is_valid_cell_ref(cell_ref):
                audit.append(
                    self._build_mapping_candidate(
                        mapping_id=mapping_id,
                        row=row,
                        sheet=sheet_name,
                        cell=cell_ref,
                        status="invalid_manual_mapping",
                        mode="manual_mapping",
                        confidence="manual",
                        score=float(selected_option.get("score", 0.0) or 0.0),
                        risk_level=str(selected_option.get("riskLevel") or "medium"),
                        requires_confirmation=False,
                        reasons=list(selected_option.get("reasons") or ["人工指定的 sheet 或 cell 无效，将回退到自动识别。"]),
                        alternative_candidates=list(candidate_item.get("alternativeCandidates", []) if isinstance(candidate_item, dict) else []),
                    )
                )
                handled_mapping_ids.discard(mapping_id)
                continue

            mapping_audit, mapping_applied = self._write_mapping_to_cell(
                workbook=workbook,
                row=row,
                mapping_id=mapping_id,
                sheet_name=sheet_name,
                cell_ref=cell_ref,
                mode="manual_mapping",
                decision_source="manual",
                write_policy=write_policy,
                candidate_option=selected_option,
            )
            audit.extend(mapping_audit)
            if mapping_applied:
                applied = True

        return TemplateFillReport(applied=applied, mode="manual_mapping", audit=audit), handled_mapping_ids

    def _apply_mapping_candidates(
        self,
        *,
        workbook: Workbook,
        rows: list[dict[str, object]],
        mapping_candidates: list[dict[str, object]],
        handled_mapping_ids: set[str],
    ) -> TemplateFillReport:
        candidate_index = {
            str(item.get("mappingId") or "").strip(): item
            for item in mapping_candidates
            if isinstance(item, dict) and str(item.get("mappingId") or "").strip()
        }
        audit: list[dict[str, object]] = []
        applied = False
        resolved_modes: list[str] = []

        for mapping_id, row in self._build_mapping_rows(rows):
            if mapping_id in handled_mapping_ids:
                continue
            candidate_item = candidate_index.get(mapping_id)
            if candidate_item is None:
                continue

            selected_option = self._select_candidate_option(candidate_item, 0)
            sheet_name = str(selected_option.get("sheet") or "").strip()
            cell_ref = str(selected_option.get("cell") or "").strip().upper()
            if bool(selected_option.get("requiresConfirmation")):
                audit.append(self._build_unconfirmed_candidate_audit(candidate_item))
                continue
            if not sheet_name or not cell_ref or sheet_name not in workbook.sheetnames or not self._is_valid_cell_ref(cell_ref):
                audit.append(
                    self._build_auto_unresolved_audit(
                        mapping_id=mapping_id,
                        row=row,
                        candidate_item=candidate_item,
                    )
                )
                continue

            mapping_audit, mapping_applied = self._write_mapping_to_cell(
                workbook=workbook,
                row=row,
                mapping_id=mapping_id,
                sheet_name=sheet_name,
                cell_ref=cell_ref,
                mode=str(candidate_item.get("mode") or "auto_mapping"),
                decision_source="auto",
                write_policy="only_empty",
                candidate_option=selected_option,
            )
            resolved_modes.append(str(candidate_item.get("mode") or "auto_mapping"))
            audit.extend(mapping_audit)
            if mapping_applied:
                applied = True

        mode = "auto_mapping"
        if resolved_modes:
            unique_modes = list(dict.fromkeys(resolved_modes))
            mode = unique_modes[0] if len(unique_modes) == 1 else "+".join(unique_modes)
        return TemplateFillReport(applied=applied, mode=mode, audit=audit)

    def _write_mapping_to_cell(
        self,
        *,
        workbook: Workbook,
        row: dict[str, object],
        mapping_id: str,
        sheet_name: str,
        cell_ref: str,
        mode: str,
        decision_source: str,
        write_policy: str,
        candidate_option: dict[str, object],
    ) -> tuple[list[dict[str, object]], bool]:
        worksheet = workbook[sheet_name]
        target_cell = worksheet[cell_ref]
        reason_summary = "；".join(str(item).strip() for item in candidate_option.get("reasons", []) if str(item).strip())
        audit_item = self._write_template_cell(
            worksheet=worksheet,
            row_index=target_cell.row,
            column_index=target_cell.column,
            value=row.get("结果"),
            mode=mode,
            row_context=row,
            field="结果",
            label=str(row.get("指标") or "").strip() or None,
            mapping_id=mapping_id,
            decision_source=decision_source,
            write_policy=write_policy,
            candidate_score=candidate_option.get("score"),
            risk_level=str(candidate_option.get("riskLevel") or "low"),
            reason_summary=reason_summary,
        )
        audit: list[dict[str, object]] = [audit_item]
        applied = audit_item["status"] in {"written", "kept_existing", "preserved_existing"}

        header_layout = self._detect_header_layout(worksheet)
        if header_layout is not None:
            result_column = self._column_index_from_ref(cell_ref)
            for header, value in row.items():
                if header == "结果":
                    continue
                column_index = header_layout.positions.get(header)
                if column_index is None or column_index == result_column:
                    continue
                side_audit = self._write_template_cell(
                    worksheet=worksheet,
                    row_index=target_cell.row,
                    column_index=column_index,
                    value=value,
                    mode=mode,
                    row_context=row,
                    field=header,
                    mapping_id=mapping_id,
                    decision_source=decision_source,
                    write_policy=write_policy,
                    candidate_score=candidate_option.get("score"),
                    risk_level=str(candidate_option.get("riskLevel") or "low"),
                    reason_summary=reason_summary,
                )
                audit.append(side_audit)
                if side_audit["status"] in {"written", "kept_existing", "preserved_existing"}:
                    applied = True

        return audit, applied

    def _fill_structured_template_sheet(
        self,
        *,
        worksheet: Worksheet,
        header_layout: HeaderLayout,
        rows: list[dict[str, object]],
    ) -> TemplateFillReport:
        """按“指标/结果”等结构化表头回填模板，并保护已有非空值。"""
        metric_col = header_layout.positions.get("指标")
        existing_metric_rows = self._read_existing_metric_rows(
            worksheet=worksheet,
            header_layout=header_layout,
        )
        metric_counts = self._count_metric_rows(rows)
        next_row = max(worksheet.max_row + 1, header_layout.row_index + 1)
        audit: list[dict[str, object]] = []
        applied = False

        for row in rows:
            metric_name = str(row.get("指标") or "").strip()
            target_row = existing_metric_rows.get(metric_name) if metric_counts.get(metric_name, 0) == 1 else None
            if target_row is None:
                # 模板里原本没有的指标会追加到后面，避免覆盖无关内容。
                target_row = next_row
                next_row += 1
                if metric_col and metric_name:
                    existing_metric_rows[metric_name] = target_row

            for header, value in row.items():
                column_index = header_layout.positions.get(header)
                if column_index is None:
                    continue
                audit_item = self._write_template_cell(
                    worksheet=worksheet,
                    row_index=target_row,
                    column_index=column_index,
                    value=value,
                    mode="header_layout",
                    row_context=row,
                    field=header,
                )
                audit.append(audit_item)
                if audit_item["status"] in {"written", "kept_existing", "preserved_existing"}:
                    applied = True

        return TemplateFillReport(applied=applied, mode="header_layout", audit=audit)

    def _preview_structured_template_sheet(
        self,
        *,
        worksheet: Worksheet,
        header_layout: HeaderLayout,
        rows: list[dict[str, object]],
    ) -> list[dict[str, object]]:
        """预览结构化模板的候选填位。"""
        result_col = header_layout.positions.get("结果")
        if result_col is None:
            return []

        metric_col = header_layout.positions.get("指标")
        existing_metric_rows = self._read_existing_metric_rows(
            worksheet=worksheet,
            header_layout=header_layout,
        )
        metric_counts = self._count_metric_rows(rows)
        next_row = max(worksheet.max_row + 1, header_layout.row_index + 1)
        candidates: list[dict[str, object]] = []

        for mapping_id, row in self._build_mapping_rows(rows):
            metric_name = str(row.get("指标") or "").strip()
            target_row = existing_metric_rows.get(metric_name) if metric_counts.get(metric_name, 0) == 1 else None
            status = "matched_existing_metric"
            confidence = "high"
            score = 0.94
            reasons = [
                f"已在工作表“{worksheet.title}”识别到结果列。",
                f"指标“{metric_name or '结果'}”命中了模板中的现有指标行。",
            ]
            alternative_candidates: list[dict[str, object]] = []
            if target_row is None:
                target_row = next_row
                next_row += 1
                if metric_col and metric_name:
                    existing_metric_rows[metric_name] = target_row
                status = "append_row"
                confidence = "medium"
                score = 0.71
                reasons = [
                    f"工作表“{worksheet.title}”已识别结果列，但没有找到指标“{metric_name or '结果'}”的现有行。",
                    "建议在模板尾部追加一行，避免覆盖已有内容。",
                ]
            else:
                alternative_candidates.append(
                    self._build_candidate_option(
                        sheet=worksheet.title,
                        cell=self._cell_coordinate(max(next_row, target_row + 1), result_col),
                        status="append_row",
                        confidence="medium",
                        score=0.63,
                        reasons=[
                            "如果不复用现有指标行，也可以改为在模板尾部追加一行。",
                        ],
                    )
                )

            candidates.append(
                self._build_mapping_candidate(
                    mapping_id=mapping_id,
                    row=row,
                    sheet=worksheet.title,
                    cell=self._cell_coordinate(target_row, result_col),
                    status=status,
                    mode="header_layout",
                    confidence=confidence,
                    score=score,
                    risk_level=self._risk_level_for_confidence(confidence),
                    requires_confirmation=confidence == "low",
                    reasons=reasons,
                    alternative_candidates=alternative_candidates,
                )
            )

        return candidates

    def _fill_label_value_template_sheet(
        self,
        *,
        worksheet: Worksheet,
        rows: list[dict[str, object]],
    ) -> TemplateFillReport:
        """按“左侧标签 + 右侧空白值”模式回填简单模板。"""
        mapping_rows = self._build_mapping_rows(rows)
        if not mapping_rows:
            return TemplateFillReport(applied=False, mode="label_value", audit=[])

        audit: list[dict[str, object]] = []
        applied = False
        matched_mapping_ids: set[str] = set()
        max_row, max_col = self._resolve_label_scan_bounds(worksheet, rows)

        for row_index in range(1, max_row + 1):
            for column_index in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row_index, column=column_index).value
                if self._is_probable_title_label(
                    worksheet=worksheet,
                    row_index=row_index,
                    column_index=column_index,
                    value=cell_value,
                ):
                    continue

                matched_mapping = self._match_label_mapping_row(
                    cell_value=cell_value,
                    mapping_rows=mapping_rows,
                    matched_mapping_ids=matched_mapping_ids,
                )
                if matched_mapping is None:
                    continue
                mapping_id, metric_row, _match_score = matched_mapping

                target_position = self._resolve_label_value_target_position(
                    worksheet=worksheet,
                    row_index=row_index,
                    column_index=column_index,
                )
                if target_position is None:
                    continue

                audit_item = self._write_template_cell(
                    worksheet=worksheet,
                    row_index=target_position[0],
                    column_index=target_position[1],
                    value=metric_row.get("结果"),
                    mode="label_value",
                    row_context=metric_row,
                    field="结果",
                    label=str(cell_value or "").strip() or str(metric_row.get("指标") or "").strip() or None,
                    mapping_id=mapping_id,
                )
                audit.append(audit_item)
                matched_mapping_ids.add(mapping_id)
                if audit_item["status"] in {"written", "kept_existing", "preserved_existing"}:
                    applied = True

        return TemplateFillReport(applied=applied, mode="label_value", audit=audit)

    def _preview_label_value_template_sheet(
        self,
        *,
        worksheet: Worksheet,
        rows: list[dict[str, object]],
    ) -> list[dict[str, object]]:
        """预览标签式模板的候选填位。"""
        mapping_rows = self._build_mapping_rows(rows)
        candidates_by_mapping_id: dict[str, list[dict[str, object]]] = {}
        max_row, max_col = self._resolve_label_scan_bounds(worksheet, rows)

        for row_index in range(1, max_row + 1):
            for column_index in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row_index, column=column_index).value
                if self._is_probable_title_label(
                    worksheet=worksheet,
                    row_index=row_index,
                    column_index=column_index,
                    value=cell_value,
                ):
                    continue

                matched_mapping = self._match_label_mapping_row(
                    cell_value=cell_value,
                    mapping_rows=mapping_rows,
                    matched_mapping_ids=set(),
                )
                if matched_mapping is None:
                    continue
                mapping_id, row, match_score = matched_mapping

                candidate_positions = self._resolve_label_value_target_positions(
                    worksheet=worksheet,
                    row_index=row_index,
                    column_index=column_index,
                )
                if not candidate_positions:
                    continue

                label_text = str(cell_value or "").strip() or str(row.get("指标") or "").strip()
                candidates_by_mapping_id.setdefault(mapping_id, []).extend(
                    self._build_candidate_option(
                        sheet=worksheet.title,
                        cell=self._cell_coordinate(item["row"], item["column"]),
                        status="matched_label",
                        confidence=self._confidence_for_label_match(float(match_score), str(item["confidence"])),
                        score=round(min(float(item["score"]) + float(match_score) / 1000, 0.99), 3),
                        reasons=[
                            f"在工作表“{worksheet.title}”中匹配到了标签“{label_text}”。",
                            f"标签同时命中指标和上下文，匹配分 {match_score}。",
                            item["reason"],
                        ],
                    )
                    for item in candidate_positions
                )

        candidates: list[dict[str, object]] = []
        for mapping_id, row in self._build_mapping_rows(rows):
            matched = self._dedupe_candidate_options(candidates_by_mapping_id.get(mapping_id, []))
            if not matched:
                candidates.append(
                    self._build_mapping_candidate(
                        mapping_id=mapping_id,
                        row=row,
                        sheet=worksheet.title,
                        cell="",
                        status="manual_review_required",
                        mode="label_value",
                        confidence="low",
                        score=0.21,
                        risk_level="high",
                        requires_confirmation=True,
                        reasons=["未找到稳定的标签值位，建议人工确认后再执行。"],
                        alternative_candidates=[],
                    )
                )
                continue

            candidates.append(
                self._build_mapping_candidate(
                    mapping_id=mapping_id,
                    row=row,
                    sheet=str(matched[0]["sheet"]),
                    cell=str(matched[0]["cell"]),
                    status=str(matched[0]["status"]),
                    mode="label_value",
                    confidence=str(matched[0]["confidence"]),
                    score=float(matched[0]["score"]),
                    risk_level=str(matched[0]["riskLevel"]),
                    requires_confirmation=bool(matched[0]["requiresConfirmation"]),
                    reasons=list(matched[0]["reasons"]),
                    alternative_candidates=matched[1:3],
                )
            )

        return candidates

    def _read_existing_metric_rows(
        self,
        *,
        worksheet: Worksheet,
        header_layout: HeaderLayout,
    ) -> dict[str, int]:
        """读取模板 sheet 中现有的“指标 -> 行号”映射。"""
        metric_col = header_layout.positions.get("指标")
        result_col = header_layout.positions.get("结果")
        if not metric_col or not result_col:
            return {}

        metric_rows: dict[str, int] = {}
        for row_index in range(header_layout.row_index + 1, worksheet.max_row + 1):
            metric_value = worksheet.cell(row=row_index, column=metric_col).value
            normalized_metric = str(metric_value).strip() if metric_value not in (None, "") else ""
            if normalized_metric:
                metric_rows[normalized_metric] = row_index
        return metric_rows

    def _count_metric_rows(self, rows: list[dict[str, object]]) -> dict[str, int]:
        """统计每个指标在结果行中的出现次数。"""
        counts: dict[str, int] = {}
        for row in rows:
            metric_name = str(row.get("指标") or "").strip()
            if metric_name:
                counts[metric_name] = counts.get(metric_name, 0) + 1
        return counts

    def _build_mapping_rows(
        self,
        rows: list[dict[str, object]],
    ) -> list[tuple[str, dict[str, object]]]:
        """为结果行生成稳定的映射 ID。"""
        return [
            (f"map_{index}", row)
            for index, row in enumerate(rows, start=1)
        ]

    def _build_mapping_candidate(
        self,
        *,
        mapping_id: str,
        row: dict[str, object],
        sheet: str,
        cell: str,
        status: str,
        mode: str,
        confidence: str,
        score: float,
        risk_level: str,
        requires_confirmation: bool,
        reasons: list[str],
        alternative_candidates: list[dict[str, object]],
    ) -> dict[str, object]:
        """构造一条候选映射记录。"""
        top_candidate = self._build_candidate_option(
            sheet=sheet,
            cell=cell,
            status=status,
            confidence=confidence,
            score=score,
            reasons=reasons,
            risk_level=risk_level,
            requires_confirmation=requires_confirmation,
        )
        return {
            "mappingId": mapping_id,
            "metric": str(row.get("指标") or "").strip() or None,
            "sheet": sheet,
            "cell": cell,
            "field": "结果",
            "status": status,
            "mode": mode,
            "confidence": confidence,
            "score": score,
            "riskLevel": risk_level,
            "requiresConfirmation": requires_confirmation,
            "reasons": reasons,
            "topCandidate": top_candidate,
            "alternativeCandidates": alternative_candidates[:2],
            "value": row.get("结果"),
            "sourceDocument": row.get("文档"),
            "sourceSheet": row.get("工作表"),
            "sourceColumn": row.get("数值字段"),
            "message": self._build_mapping_candidate_message(
                sheet=sheet,
                cell=cell,
                status=status,
            ),
        }

    def _build_candidate_option(
        self,
        *,
        sheet: str,
        cell: str,
        status: str,
        confidence: str,
        score: float,
        reasons: list[str],
        risk_level: str | None = None,
        requires_confirmation: bool | None = None,
    ) -> dict[str, object]:
        resolved_risk = risk_level or self._risk_level_for_confidence(confidence)
        resolved_requires_confirmation = (
            bool(requires_confirmation)
            if requires_confirmation is not None
            else resolved_risk == "high"
        )
        return {
            "sheet": sheet,
            "cell": cell,
            "status": status,
            "confidence": confidence,
            "score": float(score),
            "riskLevel": resolved_risk,
            "requiresConfirmation": resolved_requires_confirmation,
            "reasons": list(reasons),
            "message": self._build_mapping_candidate_message(
                sheet=sheet,
                cell=cell,
                status=status,
            ),
        }

    def _build_mapping_candidate_message(
        self,
        *,
        sheet: str,
        cell: str,
        status: str,
    ) -> str:
        """生成候选映射的人类可读说明。"""
        if status == "matched_existing_metric":
            return f"已匹配到模板工作表“{sheet}”中的现有指标行，建议写入 {cell}。"
        if status == "append_row":
            return f"模板中没有现成指标行，建议在工作表“{sheet}”追加到 {cell}。"
        if status == "matched_label":
            return f"已匹配到标签式模板，建议写入工作表“{sheet}”的 {cell}。"
        if status == "invalid_manual_mapping":
            return "人工指定的 sheet 或 cell 无效，将回退到自动识别。"
        if status == "result_sheet_fallback":
            return "未识别到稳定模板，默认将回退到新增结果表。"
        if status == "manual_mapping_skipped":
            return "该结果已被人工标记为跳过，不会写入模板。"
        if status == "requires_confirmation":
            return "该候选位置风险较高，需要人工确认后才能写入模板。"
        return "当前未能稳定定位候选单元格，建议人工确认后再执行。"

    def _risk_level_for_confidence(self, confidence: str) -> str:
        if confidence == "high":
            return "low"
        if confidence == "medium":
            return "medium"
        return "high"

    def _normalize_write_policy(self, value: object) -> str:
        normalized = str(value or "").strip()
        if normalized in {"only_empty", "allow_same", "force_overwrite"}:
            return normalized
        return "only_empty"

    def _resolve_manual_target_option(
        self,
        *,
        manual_item: dict[str, object],
        candidate_item: dict[str, object] | None,
    ) -> dict[str, object]:
        sheet_name = str(manual_item.get("sheet") or "").strip()
        cell_ref = str(manual_item.get("cell") or "").strip().upper()
        selected_index = self._coerce_candidate_index(manual_item.get("selectedCandidateIndex"))
        if sheet_name or cell_ref:
            option = self._select_candidate_option(candidate_item, selected_index)
            return {
                **option,
                "sheet": sheet_name or str(option.get("sheet") or ""),
                "cell": cell_ref or str(option.get("cell") or ""),
            }
        return self._select_candidate_option(candidate_item, selected_index)

    def _select_candidate_option(
        self,
        candidate_item: dict[str, object] | None,
        index: int,
    ) -> dict[str, object]:
        if not isinstance(candidate_item, dict):
            return {}
        options = []
        top_candidate = candidate_item.get("topCandidate")
        if isinstance(top_candidate, dict):
            options.append(top_candidate)
        options.extend(
            item for item in candidate_item.get("alternativeCandidates", [])
            if isinstance(item, dict)
        )
        if not options:
            return {
                "sheet": str(candidate_item.get("sheet") or ""),
                "cell": str(candidate_item.get("cell") or ""),
                "status": str(candidate_item.get("status") or ""),
                "confidence": str(candidate_item.get("confidence") or "low"),
                "score": float(candidate_item.get("score", 0.0) or 0.0),
                "riskLevel": str(candidate_item.get("riskLevel") or "high"),
                "requiresConfirmation": bool(candidate_item.get("requiresConfirmation")),
                "reasons": list(candidate_item.get("reasons") or []),
            }
        selected_index = max(0, min(index, len(options) - 1))
        return dict(options[selected_index])

    def _coerce_candidate_index(self, value: object) -> int:
        try:
            parsed = int(value or 0)
        except (TypeError, ValueError):
            return 0
        return max(0, parsed)

    def _dedupe_candidate_options(
        self,
        options: list[dict[str, object]],
    ) -> list[dict[str, object]]:
        deduped: list[dict[str, object]] = []
        seen_keys: set[tuple[str, str]] = set()
        for option in sorted(
            (item for item in options if isinstance(item, dict)),
            key=lambda item: (-float(item.get("score", 0.0) or 0.0), str(item.get("cell") or "")),
        ):
            key = (str(option.get("sheet") or ""), str(option.get("cell") or ""))
            if key in seen_keys:
                continue
            seen_keys.add(key)
            deduped.append(option)
        return deduped

    def _build_skipped_manual_mapping_audit(
        self,
        *,
        mapping_id: str,
        row: dict[str, object],
    ) -> dict[str, object]:
        return {
            "mappingId": mapping_id,
            "sheet": None,
            "cell": None,
            "status": "manual_mapping_skipped",
            "mode": "manual_mapping",
            "field": "结果",
            "metric": str(row.get("指标") or "").strip() or None,
            "value": row.get("结果"),
            "existingValue": None,
            "sourceDocument": row.get("文档"),
            "sourceSheet": row.get("工作表"),
            "sourceColumn": row.get("数值字段"),
            "decisionSource": "manual",
            "writePolicy": "only_empty",
            "candidateScore": None,
            "riskLevel": "low",
            "reasonSummary": "该结果已被人工显式跳过。",
            "message": "该结果已被人工标记为跳过，不会写入模板。",
        }

    def _build_unconfirmed_candidate_audit(
        self,
        candidate_item: dict[str, object],
    ) -> dict[str, object]:
        option = self._select_candidate_option(candidate_item, 0)
        return {
            "mappingId": candidate_item.get("mappingId"),
            "sheet": option.get("sheet"),
            "cell": option.get("cell"),
            "status": "requires_confirmation",
            "mode": str(candidate_item.get("mode") or "auto_mapping"),
            "field": "结果",
            "metric": candidate_item.get("metric"),
            "value": candidate_item.get("value"),
            "existingValue": None,
            "sourceDocument": candidate_item.get("sourceDocument"),
            "sourceSheet": candidate_item.get("sourceSheet"),
            "sourceColumn": candidate_item.get("sourceColumn"),
            "decisionSource": "auto",
            "writePolicy": "only_empty",
            "candidateScore": option.get("score"),
            "riskLevel": option.get("riskLevel"),
            "reasonSummary": "；".join(str(item).strip() for item in option.get("reasons", []) if str(item).strip()),
            "message": "该候选位置风险较高，需要人工确认后才能写入模板。",
        }

    def _build_auto_unresolved_audit(
        self,
        *,
        mapping_id: str,
        row: dict[str, object],
        candidate_item: dict[str, object],
    ) -> dict[str, object]:
        option = self._select_candidate_option(candidate_item, 0)
        return {
            "mappingId": mapping_id,
            "sheet": option.get("sheet"),
            "cell": option.get("cell"),
            "status": "auto_mapping_unresolved",
            "mode": str(candidate_item.get("mode") or "auto_mapping"),
            "field": "结果",
            "metric": str(row.get("指标") or "").strip() or None,
            "value": row.get("结果"),
            "existingValue": None,
            "sourceDocument": row.get("文档"),
            "sourceSheet": row.get("工作表"),
            "sourceColumn": row.get("数值字段"),
            "decisionSource": "auto",
            "writePolicy": "only_empty",
            "candidateScore": option.get("score"),
            "riskLevel": option.get("riskLevel"),
            "reasonSummary": "；".join(str(item).strip() for item in option.get("reasons", []) if str(item).strip()),
            "message": "自动候选位置无效，已回退到结果表导出。",
        }

    def _build_result_sheet_reason(self, audit_items: list[dict[str, object]]) -> str:
        if not audit_items:
            return "未识别到稳定模板，已回退为结果表导出。"
        reasons = [
            str(item.get("reasonSummary") or item.get("message") or "").strip()
            for item in audit_items
            if str(item.get("reasonSummary") or item.get("message") or "").strip()
        ]
        unique_reasons = list(dict.fromkeys(reasons))
        return "；".join(unique_reasons[:3]) or "自动回填风险较高，已改为结果表导出。"

    def _match_label_mapping_row(
        self,
        *,
        cell_value: object,
        mapping_rows: list[tuple[str, dict[str, object]]],
        matched_mapping_ids: set[str],
    ) -> tuple[str, dict[str, object], int] | None:
        """把模板标签匹配到唯一结果行。

        标签式模板经常写成“金额合计”“人数合计”“销售部合计”。
        这些标签只看 `sum` 会冲突，因此这里同时匹配指标和行上下文。
        """
        normalized_label = self._normalize_text_token(cell_value)
        if not normalized_label:
            return None

        scored_matches: list[tuple[int, int, str, dict[str, object]]] = []
        for order, (mapping_id, row) in enumerate(mapping_rows):
            if mapping_id in matched_mapping_ids:
                continue
            score = self._score_label_mapping_row(normalized_label, row)
            if score <= 0:
                continue
            scored_matches.append((score, -order, mapping_id, row))

        if not scored_matches:
            return None

        scored_matches.sort(reverse=True)
        top_score, _top_order, mapping_id, row = scored_matches[0]
        if len(scored_matches) > 1 and scored_matches[1][0] == top_score:
            return None
        return mapping_id, row, top_score

    def _score_label_mapping_row(
        self,
        normalized_label: str,
        row: dict[str, object],
    ) -> int:
        metric_name = str(row.get("指标") or "").strip()
        metric_match = self._score_metric_alias_match(normalized_label, metric_name)
        if metric_match is None:
            return 0

        context_score = self._score_label_context_match(normalized_label, row)
        if not metric_match["exact"] and context_score == 0:
            return 0

        return int(metric_match["score"]) + context_score

    def _score_metric_alias_match(self, normalized_label: str, metric_name: str) -> dict[str, object] | None:
        alias_tokens = self._metric_alias_tokens(metric_name)
        if not alias_tokens:
            return None
        if normalized_label in alias_tokens:
            return {"score": 70, "exact": True}

        contained_aliases = [
            token
            for token in alias_tokens
            if len(token) >= 2 and token in normalized_label
        ]
        if contained_aliases:
            return {"score": 46 + max(len(token) for token in contained_aliases), "exact": False}
        return None

    def _metric_alias_tokens(self, metric_name: str) -> set[str]:
        return {
            token
            for token in {
                self._normalize_text_token(metric_name),
                *(self._normalize_text_token(alias) for alias in METRIC_LABEL_ALIASES.get(metric_name, set())),
            }
            if token
        }

    def _score_label_context_match(self, normalized_label: str, row: dict[str, object]) -> int:
        score = 0
        for field_name, weight in (
            ("分组值", 18),
            ("数值字段", 16),
            ("分组字段", 8),
            ("工作表", 4),
        ):
            if self._label_contains_any_context_token(normalized_label, row.get(field_name)):
                score += weight
        return score

    def _label_contains_any_context_token(self, normalized_label: str, value: object) -> bool:
        return any(
            token in normalized_label
            for token in self._context_label_tokens(value)
        )

    def _context_label_tokens(self, value: object) -> set[str]:
        base_token = self._normalize_text_token(value)
        if not base_token:
            return set()

        tokens = {base_token}
        for alias_key, aliases in CONTEXT_LABEL_ALIASES.items():
            alias_tokens = {self._normalize_text_token(alias) for alias in aliases}
            if base_token == alias_key or base_token in alias_tokens:
                tokens.update(alias_tokens)

        return {
            token
            for token in tokens
            if len(token) >= 2 and token not in {"工作表", "sheet", "数据", "明细", "结果"}
        }

    def _confidence_for_label_match(self, match_score: float, position_confidence: str) -> str:
        if position_confidence == "low":
            return "low"
        if match_score >= 84 and position_confidence in {"high", "medium"}:
            return "high"
        if match_score >= 60:
            return "medium"
        return "low"

    def _is_probable_title_label(
        self,
        *,
        worksheet: Worksheet,
        row_index: int,
        column_index: int,
        value: object,
    ) -> bool:
        text = str(value or "").strip()
        if not text:
            return False
        if row_index > 2 or column_index > 2:
            return False
        row_non_empty = sum(
            1
            for cell in worksheet[row_index]
            if cell.value not in (None, "")
        )
        if row_non_empty > 1:
            return False
        return any(hint in text for hint in TITLE_LABEL_HINTS) and not any(
            alias in text
            for alias in ("合计", "总计", "平均", "最大", "最小", "同比", "环比", "占比")
        )

    def _resolve_label_value_target_position(
        self,
        *,
        worksheet: Worksheet,
        row_index: int,
        column_index: int,
    ) -> tuple[int, int] | None:
        """为标签式模板寻找最可能的目标值单元格。"""
        candidate_positions = self._resolve_label_value_target_positions(
            worksheet=worksheet,
            row_index=row_index,
            column_index=column_index,
        )
        if not candidate_positions:
            return None
        return candidate_positions[0]["row"], candidate_positions[0]["column"]

    def _resolve_label_value_target_positions(
        self,
        *,
        worksheet: Worksheet,
        row_index: int,
        column_index: int,
    ) -> list[dict[str, object]]:
        candidate_positions = [
            (row_index, column_index + offset, "right", offset)
            for offset in range(1, 4)
            if column_index + offset <= worksheet.max_column + 3
        ]
        candidate_positions.extend(
            (
                (row_index + offset, column_index, "down", offset)
                for offset in range(1, 3)
                if row_index + offset <= worksheet.max_row + 2
            )
        )

        candidates: list[dict[str, object]] = []
        for target_row, target_col, direction, distance in candidate_positions:
            candidate_value = worksheet.cell(row=target_row, column=target_col).value
            if self._looks_like_label(candidate_value):
                continue
            if direction == "right":
                score = 0.84 if distance == 1 else 0.66
                confidence = "medium"
            else:
                score = 0.44 if distance == 1 else 0.36
                confidence = "low"
            reason = (
                "值位紧邻标签右侧，符合常见 ESG 标签模板布局。"
                if direction == "right"
                else "值位位于标签下方，命中概率较低，建议人工确认。"
            )
            if candidate_value not in (None, ""):
                if direction == "right":
                    score -= 0.08
                    reason = "标签右侧已有值，通常仍是目标值位，但写入时会保留现有内容。"
                else:
                    score -= 0.12
                    reason = "候选单元格附近存在已有内容，命中概率下降。"
            candidates.append(
                {
                    "row": target_row,
                    "column": target_col,
                    "direction": direction,
                    "distance": distance,
                    "score": round(max(score, 0.18), 2),
                    "confidence": confidence,
                    "reason": reason,
                }
            )
        return sorted(candidates, key=lambda item: (-float(item["score"]), int(item["distance"])))

    def _replace_result_sheet(self, workbook: Workbook) -> Worksheet:
        """创建新的结果 sheet。

        如果 workbook 里已经存在旧的 `结果表`，会先替换掉，避免混入旧数据。
        """
        if RESULT_SHEET_TITLE in workbook.sheetnames:
            workbook.remove(workbook[RESULT_SHEET_TITLE])
        return workbook.create_sheet(title=RESULT_SHEET_TITLE)

    def _write_result_sheet(
        self,
        worksheet: Worksheet,
        headers: list[str],
        rows: list[dict[str, object]],
    ) -> None:
        """按普通表格方式写出结果 sheet。"""
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(header) for header in headers])

    def _write_template_cell(
        self,
        *,
        worksheet: Worksheet,
        row_index: int,
        column_index: int,
        value: object,
        mode: str,
        row_context: dict[str, object],
        field: str | None,
        label: str | None = None,
        mapping_id: str | None = None,
        decision_source: str = "auto",
        write_policy: str = "only_empty",
        candidate_score: object | None = None,
        risk_level: str = "low",
        reason_summary: str = "",
    ) -> dict[str, object]:
        """以保守策略写入模板单元格，避免覆盖已有有效值。"""
        cell = worksheet.cell(row=row_index, column=column_index)
        existing_value = cell.value
        status = "source_empty"

        merged_anchor = self._resolve_merged_anchor(
            worksheet=worksheet,
            row_index=row_index,
            column_index=column_index,
        )
        if merged_anchor == "blocked":
            status = "blocked_merged_cell"
        elif value not in (None, ""):
            if write_policy == "force_overwrite":
                cell.value = value
                status = "written"
            elif self._is_empty_like(existing_value):
                cell.value = value
                status = "written"
            elif self._cell_values_equal(existing_value, value):
                status = "kept_existing"
            else:
                status = "preserved_existing"

        return {
            "mappingId": mapping_id,
            "sheet": worksheet.title,
            "cell": cell.coordinate,
            "status": status,
            "mode": mode,
            "field": field,
            "metric": str(row_context.get("指标") or "").strip() or None,
            "value": value,
            "existingValue": existing_value,
            "sourceDocument": row_context.get("文档"),
            "sourceSheet": row_context.get("工作表"),
            "sourceColumn": row_context.get("数值字段"),
            "label": label,
            "decisionSource": decision_source,
            "writePolicy": write_policy,
            "candidateScore": float(candidate_score) if candidate_score not in (None, "") else None,
            "riskLevel": risk_level,
            "reasonSummary": reason_summary or None,
            "message": self._build_audit_message(
                status=status,
                sheet_title=worksheet.title,
                cell_coordinate=cell.coordinate,
                label=label,
            ),
        }

    def _build_audit_message(
        self,
        *,
        status: str,
        sheet_title: str,
        cell_coordinate: str,
        label: str | None,
    ) -> str:
        """生成人可读的回填审计描述。"""
        label_text = f"（标签“{label}”）" if label else ""
        if status == "written":
            return f"已在工作表“{sheet_title}”的 {cell_coordinate} 写入值{label_text}。"
        if status == "kept_existing":
            return f"工作表“{sheet_title}”的 {cell_coordinate} 已有相同值，无需重复写入{label_text}。"
        if status == "preserved_existing":
            return f"工作表“{sheet_title}”的 {cell_coordinate} 已存在不同值，已保留原值{label_text}。"
        if status == "blocked_merged_cell":
            return f"工作表“{sheet_title}”的 {cell_coordinate} 落在合并单元格区域，已阻止写入{label_text}。"
        return f"工作表“{sheet_title}”的 {cell_coordinate} 没有可写入的新值{label_text}。"

    def _score_metric_label_matches(
        self,
        worksheet: Worksheet,
        rows: list[dict[str, object]],
    ) -> int:
        """根据模板中是否出现指标标签，为候选 sheet 增加分数。"""
        mapping_rows = self._build_mapping_rows(rows)
        if not mapping_rows:
            return 0

        matched_mapping_ids: set[str] = set()
        max_row, max_col = self._resolve_label_scan_bounds(worksheet, rows)

        for row_index in range(1, max_row + 1):
            for column_index in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row_index, column=column_index).value
                if self._is_probable_title_label(
                    worksheet=worksheet,
                    row_index=row_index,
                    column_index=column_index,
                    value=cell_value,
                ):
                    continue
                matched_mapping = self._match_label_mapping_row(
                    cell_value=cell_value,
                    mapping_rows=mapping_rows,
                    matched_mapping_ids=matched_mapping_ids,
                )
                if matched_mapping is not None:
                    matched_mapping_ids.add(matched_mapping[0])

        return min(len(matched_mapping_ids), 3) * 2

    def _review_mapping_candidates(
        self,
        *,
        task: str,
        fill_plan: dict[str, object],
        mapping_candidates: list[dict[str, object]],
    ) -> dict[str, object]:
        """调用 review 模型检查候选填位计划。"""
        return self._crosscheck_service.review_mapping_plan(
            task=task,
            target_summary={
                "workbook": self._build_export_filename(list(fill_plan.get("workbookDocuments", []) or [])),
                "sheet": fill_plan.get("targetSheet"),
                "mode": fill_plan.get("mode"),
                "usedOriginalWorkbook": bool(fill_plan.get("usedOriginalWorkbook")),
                "candidateCount": len(mapping_candidates),
            },
            mapping_candidates=mapping_candidates,
        )

    def _apply_crosscheck_to_mapping_candidates(
        self,
        *,
        mapping_candidates: list[dict[str, object]],
        crosscheck: dict[str, object],
    ) -> list[dict[str, object]]:
        """把 review 的逐项风险反馈并入候选映射。"""
        reviews_by_id = {
            str(item.get("mappingId") or "").strip(): item
            for item in crosscheck.get("candidateReviews", []) or []
            if isinstance(item, dict) and str(item.get("mappingId") or "").strip()
        }
        if not reviews_by_id:
            return mapping_candidates

        reviewed_candidates: list[dict[str, object]] = []
        for candidate in mapping_candidates:
            mapping_id = str(candidate.get("mappingId") or "").strip()
            review = reviews_by_id.get(mapping_id)
            if not review:
                reviewed_candidates.append(candidate)
                continue

            reviewed_candidate = dict(candidate)
            review_risk = str(review.get("riskLevel") or "").strip() or reviewed_candidate.get("riskLevel")
            review_issue = str(review.get("issue") or "").strip()
            review_approved = bool(review.get("approved", True))
            reviewed_candidate["reviewRiskLevel"] = review_risk
            reviewed_candidate["reviewApproved"] = review_approved
            reviewed_candidate["reviewIssue"] = review_issue
            reviewed_candidate["reviewSuggestedSheet"] = str(review.get("suggestedSheet") or "").strip()
            reviewed_candidate["reviewSuggestedCell"] = str(review.get("suggestedCell") or "").strip().upper()

            if review_risk == "high" or not review_approved:
                reviewed_candidate["riskLevel"] = "high"
                reviewed_candidate["requiresConfirmation"] = True
                reviewed_candidate["confidence"] = "low"
                reviewed_candidate["status"] = "review_flagged"
            elif review_risk == "medium" and reviewed_candidate.get("riskLevel") == "low":
                reviewed_candidate["riskLevel"] = "medium"

            if review_issue:
                reasons = list(reviewed_candidate.get("reasons", []) or [])
                reviewed_candidate["reasons"] = [*reasons, f"Review 交叉检查：{review_issue}"]
                reviewed_candidate["message"] = f"Review 交叉检查提示：{review_issue}"

            reviewed_candidates.append(reviewed_candidate)

        return reviewed_candidates

    def _apply_manual_targets_to_review_candidates(
        self,
        *,
        mapping_candidates: list[dict[str, object]],
        manual_mappings: list[dict[str, object]],
    ) -> list[dict[str, object]]:
        """让 review 审查人工修改后的最终目标位置。"""
        manual_by_id = {
            str(item.get("mappingId") or "").strip(): item
            for item in manual_mappings
            if isinstance(item, dict) and str(item.get("mappingId") or "").strip()
        }
        if not manual_by_id:
            return mapping_candidates

        reviewed_candidates: list[dict[str, object]] = []
        for candidate in mapping_candidates:
            mapping_id = str(candidate.get("mappingId") or "").strip()
            manual_item = manual_by_id.get(mapping_id)
            if not manual_item:
                reviewed_candidates.append(candidate)
                continue

            reviewed_candidate = dict(candidate)
            if manual_item.get("enabled") is False:
                reviewed_candidate["status"] = "manual_mapping_skipped"
                reviewed_candidate["riskLevel"] = "low"
                reviewed_candidate["requiresConfirmation"] = False
            reviewed_candidate["sheet"] = str(manual_item.get("sheet") or reviewed_candidate.get("sheet") or "").strip()
            reviewed_candidate["cell"] = str(manual_item.get("cell") or reviewed_candidate.get("cell") or "").strip().upper()
            reviewed_candidate["writePolicy"] = self._normalize_write_policy(manual_item.get("writePolicy"))
            reviewed_candidate["decisionSource"] = "manual"
            reviewed_candidates.append(reviewed_candidate)

        return reviewed_candidates

    def _build_crosscheck_blocked_audit(
        self,
        *,
        fill_plan: dict[str, object],
        crosscheck: dict[str, object],
    ) -> dict[str, object]:
        return {
            "mappingId": None,
            "sheet": fill_plan.get("targetSheet"),
            "cell": None,
            "status": "blocked_by_crosscheck",
            "mode": str(fill_plan.get("mode") or "auto_mapping"),
            "field": "结果",
            "metric": None,
            "value": None,
            "existingValue": None,
            "sourceDocument": None,
            "sourceSheet": None,
            "sourceColumn": None,
            "decisionSource": "agent_review",
            "writePolicy": "only_empty",
            "candidateScore": None,
            "riskLevel": crosscheck.get("riskLevel"),
            "reasonSummary": "；".join(str(item) for item in crosscheck.get("issues", []) or []),
            "message": "Review 交叉检查认为当前填位计划风险较高，已阻止自动回填。",
        }

    def _resolve_label_scan_bounds(
        self,
        worksheet: Worksheet,
        rows: list[dict[str, object]],
    ) -> tuple[int, int]:
        """动态决定标签式模板扫描范围。

        小模板保持原来的轻量扫描；当结果行很多时，扫描范围随待填数量增长，
        避免 100+ / 200+ 个标签式填位落在固定前 40 行之外。
        """
        row_target = max(
            LABEL_SCAN_ROW_BASE_LIMIT,
            len(rows) * 2 + 10,
        )
        column_target = max(
            LABEL_SCAN_COLUMN_BASE_LIMIT,
            min(worksheet.max_column + 3, LABEL_SCAN_COLUMN_MAX_LIMIT),
        )
        return (
            min(max(worksheet.max_row, 1) + 3, row_target, LABEL_SCAN_ROW_MAX_LIMIT),
            min(max(worksheet.max_column, 1) + 3, column_target, LABEL_SCAN_COLUMN_MAX_LIMIT),
        )

    def _summarize_fill_audit(self, fill_report: TemplateFillReport) -> dict[str, object]:
        """汇总回填过程的写入、保留和回退情况。"""
        stats = {
            "mode": fill_report.mode,
            "written": 0,
            "keptExisting": 0,
            "preservedExisting": 0,
            "sourceEmpty": 0,
            "invalidManualMapping": 0,
            "requiresConfirmation": 0,
            "skippedManual": 0,
            "blockedMergedCell": 0,
            "blockedByCrossCheck": 0,
        }
        for item in fill_report.audit:
            status = str(item.get("status") or "").strip()
            if status == "written":
                stats["written"] += 1
            elif status == "kept_existing":
                stats["keptExisting"] += 1
            elif status == "preserved_existing":
                stats["preservedExisting"] += 1
            elif status == "source_empty":
                stats["sourceEmpty"] += 1
            elif status == "invalid_manual_mapping":
                stats["invalidManualMapping"] += 1
            elif status == "requires_confirmation":
                stats["requiresConfirmation"] += 1
            elif status == "manual_mapping_skipped":
                stats["skippedManual"] += 1
            elif status == "blocked_merged_cell":
                stats["blockedMergedCell"] += 1
            elif status == "blocked_by_crosscheck":
                stats["blockedByCrossCheck"] += 1
        return stats

    def _resolve_merged_anchor(
        self,
        *,
        worksheet: Worksheet,
        row_index: int,
        column_index: int,
    ) -> str | tuple[int, int] | None:
        for cell_range in worksheet.merged_cells.ranges:
            if row_index < cell_range.min_row or row_index > cell_range.max_row:
                continue
            if column_index < cell_range.min_col or column_index > cell_range.max_col:
                continue
            if row_index == cell_range.min_row and column_index == cell_range.min_col:
                return (row_index, column_index)
            return "blocked"
        return None

    def _merge_fill_reports(
        self,
        first: TemplateFillReport,
        second: TemplateFillReport,
    ) -> TemplateFillReport:
        """合并人工映射与自动映射的审计结果。"""
        if not first.audit:
            return second
        if not second.audit:
            return first
        merged_mode = first.mode if first.mode == second.mode else f"{first.mode}+{second.mode}"
        return TemplateFillReport(
            applied=first.applied or second.applied,
            mode=merged_mode,
            audit=[*first.audit, *second.audit],
        )

    def _looks_like_label(self, value: object) -> bool:
        """判断某个单元格内容是否更像标签而不是待填的值位。"""
        normalized = self._normalize_text_token(value)
        if not normalized:
            return False
        if normalized in {self._normalize_text_token(item) for item in HEADER_ALIASES}:
            return True
        metric_tokens = {
            self._normalize_text_token(alias)
            for aliases in METRIC_LABEL_ALIASES.values()
            for alias in aliases
        }
        if normalized in metric_tokens:
            return True
        if any(len(token) >= 2 and token in normalized for token in metric_tokens):
            return True
        return False

    def _is_empty_like(self, value: object) -> bool:
        """判断模板单元格当前值是否可视为空白。"""
        if value is None:
            return True
        text = str(value).strip()
        return self._normalize_text_token(text) in {self._normalize_text_token(item) for item in PLACEHOLDER_CELL_VALUES}

    def _cell_values_equal(self, left: object, right: object) -> bool:
        """宽松比较两个单元格值，减少无意义重复写入。"""
        if left == right:
            return True
        return self._normalize_text_token(left) == self._normalize_text_token(right)

    def _is_valid_cell_ref(self, cell_ref: str) -> bool:
        """校验 Excel 单元格引用格式。"""
        if not cell_ref:
            return False
        letters = "".join(character for character in cell_ref if character.isalpha())
        digits = "".join(character for character in cell_ref if character.isdigit())
        return bool(letters and digits and letters + digits == cell_ref)

    def _cell_coordinate(self, row_index: int, column_index: int) -> str:
        """把行列号转成 Excel 单元格坐标。"""
        column_name = ""
        current = column_index
        while current > 0:
            current, remainder = divmod(current - 1, 26)
            column_name = chr(65 + remainder) + column_name
        return f"{column_name}{row_index}"

    def _column_index_from_ref(self, cell_ref: str) -> int:
        """从 Excel 单元格坐标提取列号。"""
        letters = "".join(character for character in cell_ref if character.isalpha()).upper()
        column_index = 0
        for character in letters:
            column_index = column_index * 26 + (ord(character) - 64)
        return column_index

    def _normalize_text_token(self, value: object) -> str:
        """把文本归一化，便于做标签和占位符匹配。"""
        return "".join(
            character
            for character in str(value or "").strip().lower()
            if character.isalnum() or "\u4e00" <= character <= "\u9fff"
        )

    # 表头与命名辅助方法

    def _detect_header_layout(
        self,
        worksheet: Worksheet,
        *,
        max_scan_rows: int = 6,
    ) -> HeaderLayout | None:
        """识别 worksheet 中哪一行是真正的表头行。

        之所以会扫描前几行，是因为很多用户模板第一行是标题，第二或第三行才是表头。
        """
        scan_limit = min(max_scan_rows, worksheet.max_row)
        best_layout: HeaderLayout | None = None
        best_score = 0

        for row_index in range(1, scan_limit + 1):
            layout = self._read_row_headers(worksheet, row_index)
            if layout is None:
                continue

            score = self._score_header_layout(layout.positions)
            if score > best_score:
                best_layout = layout
                best_score = score

        if best_layout is not None:
            return best_layout

        return self._read_row_headers(worksheet, 1)

    def _read_row_headers(self, worksheet: Worksheet, row_index: int) -> HeaderLayout | None:
        """把某一行按候选表头读取出来。"""
        if row_index < 1 or row_index > worksheet.max_row:
            return None

        positions: dict[str, int] = {}
        headers: list[str] = []
        for cell in worksheet[row_index]:
            if cell.value in (None, ""):
                continue
            header = str(cell.value).strip()
            if not header:
                continue
            headers.append(header)
            if header not in positions:
                positions[header] = cell.column
            for canonical, aliases in HEADER_ALIASES.items():
                if header.lower() in {alias.lower() for alias in aliases} and canonical not in positions:
                    positions[canonical] = cell.column

        if not positions:
            return None

        return HeaderLayout(row_index=row_index, headers=headers, positions=positions)

    def _score_header_layout(self, positions: dict[str, int]) -> int:
        """给候选表头行打分，分数越高越像结果表表头。"""
        score = 0
        if "指标" in positions:
            score += 4
        if "结果" in positions:
            score += 4
        if "文档" in positions:
            score += 1
        if "工作表" in positions:
            score += 1
        if "数值字段" in positions:
            score += 1
        return score

    def _make_unique_sheet_name(self, workbook: Workbook, raw_title: str) -> str:
        """生成合法且不重复的 Excel sheet 名称。"""
        base_title = (raw_title or "Sheet").strip()[:31] or "Sheet"
        candidate = base_title
        suffix = 2
        while candidate in workbook.sheetnames:
            suffix_text = f"_{suffix}"
            candidate = f"{base_title[: 31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        return candidate

    def _build_export_filename(self, documents: list[Any]) -> str:
        """根据原始 Excel 文件名生成下载文件名。"""
        for document in documents:
            if document.type != "excel":
                continue
            base_name = str(document.name or "workflow_result.xlsx").rsplit(".", 1)[0]
            return f"{base_name}_filled.xlsx"
        return "workflow_result_filled.xlsx"

    def _build_template_fill_summary(
        self,
        *,
        sheet_title: str,
        used_original_workbook: bool,
    ) -> str:
        """生成“已回填模板 sheet”场景下的导出说明。"""
        if used_original_workbook:
            return f"已回填到原始工作簿中的模板工作表“{sheet_title}”。"
        return f"已回填到模板工作表“{sheet_title}”。"

    def _build_result_sheet_summary(
        self,
        *,
        sheet_title: str,
        used_original_workbook: bool,
    ) -> str:
        """生成“新建结果 sheet”场景下的导出说明。"""
        if used_original_workbook:
            return f"已在原始工作簿中新增结果工作表“{sheet_title}”。"
        return f"已生成新的结果工作表“{sheet_title}”。"
