"""Excel 回填校验技能。"""

from __future__ import annotations

import base64
import binascii
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.services.workflow.skills.base import SkillExecutionContext, WorkflowSkill


class FillValidatorSkill(WorkflowSkill):
    """写入后重新读取导出的 workbook，确认目标单元格结果一致。"""

    name = "fill_validator"
    title = "校验填表结果"
    description = "回填完成后重新读取导出的 Excel，校验已写入单元格和审计记录是否一致。"
    input_hint = "table_filler 的导出文件和 fillAudit"
    output_hint = "逐单元格校验结果和通过/失败统计"
    tags = ("excel", "spreadsheet", "fill", "validate")

    def execute(self, context: SkillExecutionContext) -> dict[str, object]:
        filler_result = context.previous_results.get("table_filler", {})
        export_files = [
            item
            for item in filler_result.get("exportFiles", []) or []
            if isinstance(item, dict) and self._is_excel_export_file(item)
        ]
        fill_audit = [item for item in filler_result.get("fillAudit", []) or [] if isinstance(item, dict)]
        if not export_files:
            return {
                "summary": "当前没有可校验的 Excel 导出文件。",
                "validationStats": {"checked": 0, "passed": 0, "failed": 0, "skipped": len(fill_audit)},
                "validations": [],
            }

        content_base64 = str(export_files[0].get("contentBase64") or "").strip()
        if not content_base64:
            return {
                "summary": "当前导出文件未内联内容，无法在本步骤重新读取校验。",
                "validationStats": {"checked": 0, "passed": 0, "failed": 0, "skipped": len(fill_audit)},
                "validations": [],
            }

        try:
            workbook_bytes = base64.b64decode(content_base64, validate=True)
            workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
        except (binascii.Error, ValueError, OSError, KeyError) as exc:
            return {
                "summary": f"导出文件内容无法读取，已跳过回填校验：{type(exc).__name__}。",
                "validationStats": {
                    "checked": 0,
                    "passed": 0,
                    "failed": 0,
                    "skipped": len(fill_audit),
                    "readError": True,
                },
                "validations": [],
            }

        validations: list[dict[str, object]] = []
        for item in fill_audit:
            if str(item.get("field") or "") != "结果":
                continue
            if str(item.get("status") or "") not in {"written", "kept_existing"}:
                continue
            sheet_name = str(item.get("sheet") or "").strip()
            cell_ref = str(item.get("cell") or "").strip()
            if not sheet_name or not cell_ref or sheet_name not in workbook.sheetnames:
                validations.append(self._build_validation_item(item=item, actual=None, passed=False, reason="目标 sheet 或 cell 不存在。"))
                continue
            actual = workbook[sheet_name][cell_ref].value
            passed = self._cell_values_equal(actual, item.get("value"))
            validations.append(
                self._build_validation_item(
                    item=item,
                    actual=actual,
                    passed=passed,
                    reason="实际值与预期一致。" if passed else "实际值与预期不一致。",
                )
            )
        workbook.close()

        passed_count = sum(1 for item in validations if bool(item.get("passed")))
        failed_count = sum(1 for item in validations if not bool(item.get("passed")))
        skipped_count = len(fill_audit) - len(validations)
        summary = f"已校验 {len(validations)} 个回填单元格，{passed_count} 个通过，{failed_count} 个失败。"
        return {
            "summary": summary,
            "validationStats": {
                "checked": len(validations),
                "passed": passed_count,
                "failed": failed_count,
                "skipped": skipped_count,
            },
            "validations": validations,
        }

    def _build_validation_item(
        self,
        *,
        item: dict[str, object],
        actual: object,
        passed: bool,
        reason: str,
    ) -> dict[str, object]:
        return {
            "mappingId": item.get("mappingId"),
            "sheet": item.get("sheet"),
            "cell": item.get("cell"),
            "expected": item.get("value"),
            "actual": actual,
            "passed": passed,
            "reason": reason,
        }

    def _cell_values_equal(self, left: Any, right: Any) -> bool:
        if left == right:
            return True
        return self._normalize_cell_value(left) == self._normalize_cell_value(right)

    def _normalize_cell_value(self, value: Any) -> str:
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value if value is not None else "").strip().lower()

    def _is_excel_export_file(self, item: dict[str, object]) -> bool:
        mime_type = str(item.get("mimeType") or "").strip().lower()
        filename = str(item.get("filename") or "").strip().lower()
        if "spreadsheetml" in mime_type or "excel" in mime_type:
            return True
        if filename.endswith((".xlsx", ".xlsm")):
            return True
        return not mime_type and not filename
